import openpyxl
import json
import os
import pickle
from typing import List, Dict
from pathlib import Path

CACHE_FILE = ".products_cache.pkl"

def load_excel_products() -> List[Dict]:
    """从Excel文件加载所有产品（带缓存）"""
    all_products = []
    
    excel_file = "./attached_assets/Malaysia_Financial_Products_Master_Clean_1762842694551.xlsx"
    
    # 检查缓存
    if os.path.exists(CACHE_FILE):
        excel_mtime = os.path.getmtime(excel_file)
        cache_mtime = os.path.getmtime(CACHE_FILE)
        if cache_mtime > excel_mtime:
            try:
                with open(CACHE_FILE, 'rb') as f:
                    cached_data = pickle.load(f)
                    if cached_data.get('version') == '1.0':
                        return cached_data.get('excel_products', [])
            except:
                pass
    
    # 如果没有缓存，读取Excel
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary_Report":
                continue
                
            sheet = wb[sheet_name]
            
            # 读取表头
            headers = []
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell_value = sheet.cell(1, col).value
                headers.append(str(cell_value).strip() if cell_value else f"Col{col}")
            
            # 读取所有数据行
            for row in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = sheet.cell(row, col_idx).value
                    row_data[header] = str(cell_value).strip()[:200] if cell_value else ""
                
                # 只添加非空行
                if any(row_data.values()):
                    row_data['_source'] = f"Excel-{sheet_name}"
                    all_products.append(row_data)
        
        wb.close()
        
        # 保存缓存
        try:
            with open(CACHE_FILE, 'wb') as f:
                pickle.dump({'version': '1.0', 'excel_products': all_products}, f)
        except:
            pass
            
    except Exception as e:
        print(f"Excel读取错误: {e}")
    
    return all_products

def load_json_products() -> List[Dict]:
    """从JSON文件加载所有产品"""
    all_products = []
    
    banks_dir = "data/banks"
    
    if not os.path.exists(banks_dir):
        return all_products
    
    for filename in os.listdir(banks_dir):
        if not filename.endswith('.json'):
            continue
            
        filepath = os.path.join(banks_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                products = []
                if isinstance(data, list):
                    products = data
                elif isinstance(data, dict):
                    if 'products' in data:
                        products = data['products']
                    elif 'data' in data:
                        products = data['data']
                
                for product in products:
                    product['_source'] = f"JSON-{filename}"
                    all_products.append(product)
                    
        except Exception as e:
            print(f"JSON读取错误 {filename}: {e}")
    
    return all_products

def get_all_products():
    """获取所有产品数据"""
    excel_products = load_excel_products()
    json_products = load_json_products()
    
    return {
        'excel': excel_products,
        'json': json_products,
        'total': len(excel_products) + len(json_products)
    }

if __name__ == "__main__":
    result = get_all_products()
    print(f"Excel产品: {len(result['excel'])}")
    print(f"JSON产品: {len(result['json'])}")
    print(f"总计: {result['total']}")
