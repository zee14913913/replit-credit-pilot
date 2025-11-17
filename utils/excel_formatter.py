"""
CreditPilot Professional Excel Formatter
使用官方配色方案（主粉色#FFB6C1 + 深棕色#3E2723）美化Excel报告

功能：
- 13项专业格式化要求
- 统一的CreditPilot配色方案
- 自动列宽、行高、边框、颜色
- 冻结窗格、筛选器、打印设置
"""

import json
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, 
        numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，Excel格式化功能不可用")


class ExcelFormatter:
    """专业级Excel格式化工具类"""
    
    def __init__(self, colors_config_path: str = "config/colors.json"):
        """
        初始化Excel格式化器
        
        Args:
            colors_config_path: 色彩配置文件路径
        """
        self.load_colors(colors_config_path)
        
    def load_colors(self, config_path: str):
        """加载CreditPilot官方配色方案"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                excel_colors = config['creditpilot_theme']['excel_formatting']
                
                self.COLORS = {
                    'header_bg': excel_colors['header_row']['background'],
                    'header_text': excel_colors['header_row']['text'],
                    'odd_row': excel_colors['data_rows']['odd_row'],
                    'even_row': excel_colors['data_rows']['even_row'],
                    'owners_expenses': excel_colors['category_colors']['owners_expenses'],
                    'gz_expenses': excel_colors['category_colors']['gz_expenses'],
                    'suppliers': excel_colors['category_colors']['suppliers'],
                    'payments': excel_colors['category_colors']['payments'],
                    'outstanding_bg': excel_colors['category_colors']['outstanding_balance']['background'],
                    'outstanding_text': excel_colors['category_colors']['outstanding_balance']['text'],
                    'border_standard': excel_colors['borders']['standard'],
                    'border_light': excel_colors['borders']['light'][:7],
                    'negative_amount': excel_colors['special_formats']['negative_amount'],
                    'overdue': excel_colors['special_formats']['overdue_date']
                }
        except FileNotFoundError:
            print(f"⚠️ 配色文件未找到: {config_path}，使用默认配色")
            self._load_default_colors()
    
    def _load_default_colors(self):
        """默认CreditPilot配色方案"""
        self.COLORS = {
            'header_bg': '#3E2723',
            'header_text': '#FFFFFF',
            'odd_row': '#FFFFFF',
            'even_row': '#FAF8F7',
            'owners_expenses': '#FFE4E1',
            'gz_expenses': '#F5E6E8',
            'suppliers': '#F8E8E6',
            'payments': '#F3E5F5',
            'outstanding_bg': '#FFB6C1',
            'outstanding_text': '#3E2723',
            'border_standard': '#3E2723',
            'border_light': '#3E2723',
            'negative_amount': '#D32F2F',
            'overdue': '#FFE4E1'
        }
    
    def format_worksheet(self, 
                        ws: Worksheet, 
                        sheet_type: str,
                        customer_name: str = "Cheok Jun Yoon") -> None:
        """
        格式化整个工作表
        
        Args:
            ws: openpyxl工作表对象
            sheet_type: 工作表类型 (summary/transactions/categories/errors)
            customer_name: 客户名称
        """
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl未安装，跳过格式化")
            return
        
        self._set_column_widths(ws, sheet_type)
        self._format_header_row(ws)
        self._format_data_rows(ws, sheet_type)
        self._apply_borders(ws)
        self._set_row_heights(ws)
        self._freeze_panes(ws)
        self._add_filters(ws)
        self._set_page_setup(ws, customer_name)
    
    def _set_column_widths(self, ws: Worksheet, sheet_type: str):
        """设置列宽（要求1）"""
        column_widths = {
            'summary': {
                'A': 20,  # bank_name
                'B': 22,  # card_number
                'C': 18,  # statement_date
                'D': 18,  # due_date
                'E': 15,  # previous_balance
                'F': 15,  # total_expenses
                'G': 15,  # total_payments
                'H': 15,  # outstanding_balance
                'I': 15,  # minimum_payment
                'J': 15   # owners_balance
            },
            'transactions': {
                'A': 20,  # bank_name
                'B': 22,  # card_number
                'C': 18,  # transaction_date
                'D': 40,  # description
                'E': 15,  # amount
                'F': 20,  # category
                'G': 15,  # supplier_fee
                'H': 18   # statement_date
            },
            'categories': {
                'A': 25,  # category
                'B': 15,  # count
                'C': 18,  # total_amount
                'D': 15,  # percentage
                'E': 40   # description
            },
            'errors': {
                'A': 20,  # bank_name
                'B': 22,  # card_number
                'C': 18,  # statement_date
                'D': 40,  # error_message
                'E': 25   # details
            }
        }
        
        widths = column_widths.get(sheet_type, column_widths['summary'])
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _format_header_row(self, ws: Worksheet):
        """格式化标题行（要求2、4、5、6）"""
        if ws.max_row < 1:
            return
        
        header_font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color=self.COLORS['header_text'].replace('#', '')
        )
        
        header_fill = PatternFill(
            start_color=self.COLORS['header_bg'].replace('#', ''),
            end_color=self.COLORS['header_bg'].replace('#', ''),
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=False
        )
        
        thick_border = Border(
            left=Side(style='thin', color=self.COLORS['border_standard'].replace('#', '')),
            right=Side(style='thin', color=self.COLORS['border_standard'].replace('#', '')),
            top=Side(style='thin', color=self.COLORS['border_standard'].replace('#', '')),
            bottom=Side(style='medium', color=self.COLORS['border_standard'].replace('#', ''))
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thick_border
        
        ws.row_dimensions[1].height = 30
    
    def _format_data_rows(self, ws: Worksheet, sheet_type: str):
        """格式化数据行（要求2、5、7、11）"""
        if ws.max_row < 2:
            return
        
        data_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', size=11, bold=True)
        
        for row_idx in range(2, ws.max_row + 1):
            row_bg_color = self.COLORS['odd_row'] if row_idx % 2 == 0 else self.COLORS['even_row']
            
            fill = PatternFill(
                start_color=row_bg_color.replace('#', ''),
                end_color=row_bg_color.replace('#', ''),
                fill_type='solid'
            )
            
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.font = data_font
                cell.fill = fill
                
                col_letter = get_column_letter(col_idx)
                alignment = self._get_column_alignment(col_letter, sheet_type)
                cell.alignment = alignment
                
                if cell.value and isinstance(cell.value, (int, float)):
                    self._format_number_cell(cell)
                elif cell.value and isinstance(cell.value, str) and 'date' in str(ws.cell(1, col_idx).value).lower():
                    cell.number_format = 'DD/MM/YYYY'
            
            ws.row_dimensions[row_idx].height = 25
    
    def _get_column_alignment(self, col_letter: str, sheet_type: str) -> Alignment:
        """获取列对齐方式"""
        amount_columns = {'E', 'F', 'G', 'H', 'I', 'J', 'C'}
        date_columns = {'C', 'D', 'H'}
        
        if col_letter in amount_columns and sheet_type in ['summary', 'transactions', 'categories']:
            return Alignment(horizontal='right', vertical='center', wrap_text=False)
        elif col_letter in date_columns:
            return Alignment(horizontal='center', vertical='center', wrap_text=False)
        else:
            return Alignment(horizontal='left', vertical='center', wrap_text=False)
    
    def _format_number_cell(self, cell):
        """格式化数字单元格（要求7、11）"""
        if cell.value < 0:
            cell.font = Font(name='Calibri', size=11, color=self.COLORS['negative_amount'].replace('#', ''))
            cell.number_format = '"RM" (#,##0.00)'
        else:
            cell.number_format = '"RM" #,##0.00'
    
    def _apply_borders(self, ws: Worksheet):
        """应用边框（要求4）"""
        thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border_light'].replace('#', '')),
            right=Side(style='thin', color=self.COLORS['border_light'].replace('#', '')),
            top=Side(style='thin', color=self.COLORS['border_light'].replace('#', '')),
            bottom=Side(style='thin', color=self.COLORS['border_light'].replace('#', ''))
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.border == Border():
                    cell.border = thin_border
    
    def _set_row_heights(self, ws: Worksheet):
        """设置行高（要求2）"""
        for row_idx in range(1, ws.max_row + 1):
            if row_idx == 1:
                ws.row_dimensions[row_idx].height = 30
            else:
                ws.row_dimensions[row_idx].height = 25
    
    def _freeze_panes(self, ws: Worksheet):
        """冻结窗格（要求8）"""
        ws.freeze_panes = 'A2'
    
    def _add_filters(self, ws: Worksheet):
        """添加筛选器（要求9）"""
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    
    def _set_page_setup(self, ws: Worksheet, customer_name: str):
        """设置页面布局（要求10）"""
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 1.5 / 2.54
        ws.page_margins.right = 1.5 / 2.54
        ws.page_margins.top = 1.5 / 2.54
        ws.page_margins.bottom = 1.5 / 2.54
        
        ws.oddHeader.center.text = f"&B{customer_name} - Credit Card Settlement Report"
        ws.oddHeader.right.text = f"&D"
        ws.oddFooter.center.text = "Page &P of &N"
    
    def format_category_summary_row(self, ws: Worksheet, row_idx: int, category: str):
        """格式化分类汇总行（要求12）"""
        category_colors = {
            'Owners Expenses': self.COLORS['owners_expenses'],
            'GZ Expenses': self.COLORS['gz_expenses'],
            'Suppliers': self.COLORS['suppliers'],
            'Owners Payment': self.COLORS['payments'],
            'GZ Payment': self.COLORS['payments'],
            'Outstanding Balance': self.COLORS['outstanding_bg']
        }
        
        bg_color = category_colors.get(category, self.COLORS['even_row'])
        
        fill = PatternFill(
            start_color=bg_color.replace('#', ''),
            end_color=bg_color.replace('#', ''),
            fill_type='solid'
        )
        
        bold_font = Font(name='Calibri', size=11, bold=True)
        if category == 'Outstanding Balance':
            bold_font = Font(
                name='Calibri', 
                size=11, 
                bold=True,
                color=self.COLORS['outstanding_text'].replace('#', '')
            )
        
        thick_top_border = Border(
            left=Side(style='thin', color=self.COLORS['border_standard'].replace('#', '')),
            right=Side(style='thin', color=self.COLORS['border_standard'].replace('#', '')),
            top=Side(style='medium', color=self.COLORS['border_standard'].replace('#', '')),
            bottom=Side(style='thin', color=self.COLORS['border_standard'].replace('#', ''))
        )
        
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = bold_font
            cell.border = thick_top_border
            
            if cell.column > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
    
    def add_transaction_icons(self, ws: Worksheet, description_col: str = 'D'):
        """添加交易图标（要求11）"""
        for row_idx in range(2, ws.max_row + 1):
            category_cell = ws[f'F{row_idx}']
            desc_cell = ws[f'{description_col}{row_idx}']
            
            if category_cell.value == 'Suppliers' and desc_cell.value:
                desc_cell.value = f"🏪 {desc_cell.value}"
            elif category_cell.value in ['GZ Expenses', 'GZ Payment'] and desc_cell.value:
                desc_cell.value = f"💼 {desc_cell.value}"


def create_formatted_excel(data: Dict[str, Any], 
                          output_path: str,
                          customer_name: str = "Cheok Jun Yoon") -> str:
    """
    创建专业格式化的Excel报告
    
    Args:
        data: 处理结果数据
        output_path: 输出文件路径
        customer_name: 客户名称
    
    Returns:
        生成的文件路径
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl未安装，无法生成格式化Excel报告")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    formatter = ExcelFormatter()
    
    _create_summary_sheet(wb, data, formatter, customer_name)
    _create_transactions_sheet(wb, data, formatter, customer_name)
    _create_categories_sheet(wb, data, formatter, customer_name)
    _create_errors_sheet(wb, data, formatter, customer_name)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_path = output_path.replace('.xlsx', f'_{timestamp}.xlsx')
    final_path = final_path.replace('settlement_report', f'{customer_name.replace(" ", "")}_Settlement')
    
    wb.save(final_path)
    return final_path


def _create_summary_sheet(wb: Workbook, data: Dict, formatter: ExcelFormatter, customer_name: str):
    """创建账单汇总工作表"""
    ws = wb.create_sheet("账单汇总")
    
    headers = ['银行', '卡号', '账单日期', '到期日期', '上期余额', '本期消费', 
               '本期还款', 'Outstanding Balance', '最低还款', 'Owners余额']
    ws.append(headers)
    
    for statement in data.get('statements', []):
        row = [
            statement.get('bank_name', ''),
            statement.get('card_number', ''),
            statement.get('statement_date', ''),
            statement.get('due_date', ''),
            statement.get('previous_balance', 0),
            statement.get('total_expenses', 0),
            statement.get('total_payments', 0),
            statement.get('outstanding_balance', 0),
            statement.get('minimum_payment', 0),
            statement.get('owners_balance', 0)
        ]
        ws.append(row)
    
    formatter.format_worksheet(ws, 'summary', customer_name)


def _create_transactions_sheet(wb: Workbook, data: Dict, formatter: ExcelFormatter, customer_name: str):
    """创建交易明细工作表"""
    ws = wb.create_sheet("交易明细")
    
    headers = ['银行', '卡号', '交易日期', '描述', '金额', '分类', '供应商手续费', '账单日期']
    ws.append(headers)
    
    for transaction in data.get('transactions', []):
        row = [
            transaction.get('bank_name', ''),
            transaction.get('card_number', ''),
            transaction.get('transaction_date', ''),
            transaction.get('description', ''),
            transaction.get('amount', 0),
            transaction.get('category', ''),
            transaction.get('supplier_fee', 0),
            transaction.get('statement_date', '')
        ]
        ws.append(row)
    
    formatter.format_worksheet(ws, 'transactions', customer_name)
    formatter.add_transaction_icons(ws, 'D')


def _create_categories_sheet(wb: Workbook, data: Dict, formatter: ExcelFormatter, customer_name: str):
    """创建分类汇总工作表"""
    ws = wb.create_sheet("分类汇总")
    
    headers = ['分类', '交易数量', '总金额', '占比%', '说明']
    ws.append(headers)
    
    categories = data.get('category_summary', {})
    for category, info in categories.items():
        row = [
            category,
            info.get('count', 0),
            info.get('total', 0),
            info.get('percentage', 0),
            info.get('description', '')
        ]
        ws.append(row)
        formatter.format_category_summary_row(ws, ws.max_row, category)
    
    formatter.format_worksheet(ws, 'categories', customer_name)


def _create_errors_sheet(wb: Workbook, data: Dict, formatter: ExcelFormatter, customer_name: str):
    """创建错误记录工作表"""
    ws = wb.create_sheet("错误记录")
    
    headers = ['银行', '卡号', '账单日期', '错误信息', '详情']
    ws.append(headers)
    
    for error in data.get('errors', []):
        row = [
            error.get('bank_name', ''),
            error.get('card_number', ''),
            error.get('statement_date', ''),
            error.get('error_message', ''),
            error.get('details', '')
        ]
        ws.append(row)
    
    formatter.format_worksheet(ws, 'errors', customer_name)
