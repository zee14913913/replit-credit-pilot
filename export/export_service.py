"""
Export Service - Data Export Functionality
Supports Excel, CSV, and Enhanced PDF exports
"""

import csv
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db.database import get_db

class ExportService:
    """Service for exporting transaction data in various formats"""
    
    EXPORT_DIR = 'static/exports'
    
    def __init__(self):
        os.makedirs(self.EXPORT_DIR, exist_ok=True)
    
    def export_to_excel(self, customer_id: int, filters: Optional[Dict] = None) -> str:
        """Export transactions to Excel with professional formatting"""
        
        # Fetch data
        transactions = self._get_filtered_transactions(customer_id, filters)
        customer = self._get_customer_info(customer_id)
        spending_summary = self._calculate_spending_summary(transactions)
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Transactions
        ws_trans = wb.active
        ws_trans.title = "Transactions"
        self._create_transactions_sheet(ws_trans, transactions, customer)
        
        # Sheet 2: Summary
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, spending_summary, customer)
        
        # Sheet 3: Monthly Trends
        ws_trends = wb.create_sheet("Monthly Trends")
        self._create_trends_sheet(ws_trends, transactions, customer)
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"transactions_{customer['name']}_{timestamp}.xlsx"
        filepath = os.path.join(self.EXPORT_DIR, filename)
        wb.save(filepath)
        
        # Log export
        self._log_export(customer_id, 'transactions', 'excel', filepath, len(transactions), filters)
        
        return filepath
    
    def export_to_csv(self, customer_id: int, filters: Optional[Dict] = None) -> str:
        """Export transactions to CSV format"""
        
        transactions = self._get_filtered_transactions(customer_id, filters)
        customer = self._get_customer_info(customer_id)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"transactions_{customer['name']}_{timestamp}.csv"
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Date', 'Description', 'Amount (RM)', 'Category', 'Notes', 'Tags']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for trans in transactions:
                writer.writerow({
                    'Date': trans['transaction_date'],
                    'Description': trans['description'],
                    'Amount (RM)': f"{trans['amount']:.2f}",
                    'Category': trans['category'],
                    'Notes': trans.get('notes', ''),
                    'Tags': trans.get('tags', '')
                })
        
        self._log_export(customer_id, 'transactions', 'csv', filepath, len(transactions), filters)
        
        return filepath
    
    def _create_transactions_sheet(self, ws, transactions: List[Dict], customer: Dict):
        """Create professionally formatted transactions sheet"""
        
        # Header styling
        header_fill = PatternFill(start_color="1FAA59", end_color="1FAA59", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"Transaction Report - {customer['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="F5E6C8")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws['A2'].font = Font(size=10, color="94A3B8")
        ws.merge_cells('A2:F2')
        
        # Headers
        headers = ['Date', 'Description', 'Amount (RM)', 'Category', 'Notes', 'Tags']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_num, trans in enumerate(transactions, 5):
            ws.cell(row=row_num, column=1, value=trans['transaction_date']).border = border
            ws.cell(row=row_num, column=2, value=trans['description']).border = border
            
            amount_cell = ws.cell(row=row_num, column=3, value=trans['amount'])
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = border
            
            ws.cell(row=row_num, column=4, value=trans['category']).border = border
            ws.cell(row=row_num, column=5, value=trans.get('notes', '')).border = border
            ws.cell(row=row_num, column=6, value=trans.get('tags', '')).border = border
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['E'].width = 30
    
    def _create_summary_sheet(self, ws, summary: Dict, customer: Dict):
        """Create spending summary sheet"""
        
        header_fill = PatternFill(start_color="F5E6C8", end_color="F5E6C8", fill_type="solid")
        header_font = Font(bold=True, size=12)
        
        ws['A1'] = f"Spending Summary - {customer['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="1FAA59")
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Category'
        ws['B3'] = 'Amount (RM)'
        ws['C3'] = 'Transactions'
        ws['D3'] = 'Percentage'
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = Alignment(horizontal='center')
        
        total_amount = sum(data['total'] for data in summary.values())
        
        row = 4
        for category, data in sorted(summary.items(), key=lambda x: x[1]['total'], reverse=True):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=data['total']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=data['count'])
            percentage = (data['total'] / total_amount * 100) if total_amount > 0 else 0
            ws.cell(row=row, column=4, value=percentage / 100).number_format = '0.0%'
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_amount).number_format = '#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=sum(data['count'] for data in summary.values()))
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=1.0).number_format = '0.0%'
        ws.cell(row=row, column=4).font = Font(bold=True)
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_trends_sheet(self, ws, transactions: List[Dict], customer: Dict):
        """Create monthly trends analysis sheet"""
        
        # Group by month
        monthly_data = {}
        for trans in transactions:
            try:
                date = datetime.strptime(trans['transaction_date'], '%Y-%m-%d')
                month_key = date.strftime('%Y-%m')
                if month_key not in monthly_data:
                    monthly_data[month_key] = {'total': 0, 'count': 0}
                monthly_data[month_key]['total'] += trans['amount']
                monthly_data[month_key]['count'] += 1
            except:
                continue
        
        ws['A1'] = f"Monthly Spending Trends - {customer['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="1FAA59")
        ws.merge_cells('A1:C1')
        
        ws['A3'] = 'Month'
        ws['B3'] = 'Total Spending (RM)'
        ws['C3'] = 'Transactions'
        
        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="F5E6C8", end_color="F5E6C8", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal='center')
        
        row = 4
        for month in sorted(monthly_data.keys(), reverse=True):
            data = monthly_data[month]
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=data['total']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=data['count'])
            row += 1
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 20
    
    def _get_filtered_transactions(self, customer_id: int, filters: Optional[Dict]) -> List[Dict]:
        """Fetch transactions with optional filters"""
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            query = '''
                SELECT t.*, GROUP_CONCAT(tg.tag_name, ', ') as tags
                FROM transactions t
                INNER JOIN statements s ON t.statement_id = s.id
                INNER JOIN credit_cards cc ON s.card_id = cc.id
                LEFT JOIN transaction_tags tt ON t.id = tt.transaction_id
                LEFT JOIN tags tg ON tt.tag_id = tg.id
                WHERE cc.customer_id = ? AND s.is_confirmed = 1
            '''
            
            params = [customer_id]
            
            if filters:
                if filters.get('start_date'):
                    query += ' AND t.transaction_date >= ?'
                    params.append(filters['start_date'])
                if filters.get('end_date'):
                    query += ' AND t.transaction_date <= ?'
                    params.append(filters['end_date'])
                if filters.get('category'):
                    query += ' AND t.category = ?'
                    params.append(filters['category'])
                if filters.get('min_amount'):
                    query += ' AND t.amount >= ?'
                    params.append(filters['min_amount'])
                if filters.get('max_amount'):
                    query += ' AND t.amount <= ?'
                    params.append(filters['max_amount'])
            
            query += ' GROUP BY t.id ORDER BY t.transaction_date DESC'
            
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
    
    def _get_customer_info(self, customer_id: int) -> Dict:
        """Get customer information"""
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM customers WHERE id = ?', (customer_id,))
            row = cursor.fetchone()
            return dict(row) if row else {}
    
    def _calculate_spending_summary(self, transactions: List[Dict]) -> Dict:
        """Calculate spending summary by category"""
        summary = {}
        for trans in transactions:
            category = trans['category'] or 'Uncategorized'
            if category not in summary:
                summary[category] = {'total': 0, 'count': 0}
            summary[category]['total'] += trans['amount']
            summary[category]['count'] += 1
        return summary
    
    def _log_export(self, customer_id: int, export_type: str, export_format: str, 
                    file_path: str, record_count: int, filters: Optional[Dict]):
        """Log export operation to database"""
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO export_history 
                (customer_id, export_type, export_format, file_path, record_count, filters_applied)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (customer_id, export_type, export_format, file_path, record_count, 
                  str(filters) if filters else None))
            conn.commit()
