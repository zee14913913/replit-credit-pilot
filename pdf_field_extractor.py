#!/usr/bin/env python3
"""
信用卡账单字段提取系统 - 多银行支持版
=======================================
严格遵守规则：
1. 必须从PDF真实解析：Statement Date、Due Date、Statement Total、Minimum Payment
2. 绝对禁止使用公式计算、固定值、估算
3. 无法识别时返回NULL并记录异常
4. 每张账单的值必须唯一、可追溯
"""

import os
import sqlite3
import pdfplumber
import re
from datetime import datetime
from decimal import Decimal
from typing import Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class BankParser:
    """银行账单解析基类"""
    
    def __init__(self, bank_name: str):
        self.bank_name = bank_name
    
    def parse(self, pdf_path: str) -> Dict:
        """解析PDF，返回4个关键字段"""
        return {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }

class HongLeongBankParser(BankParser):
    """Hong Leong Bank / HLB 专用解析器"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # 1. Statement Date - "Statement Date 07 SEP 2024"
                date_match = re.search(r'Statement\s+Date\s+(\d{2}\s+[A-Z]{3}\s+\d{4})', text)
                if date_match:
                    date_str = date_match.group(1)
                    try:
                        result['statement_date'] = datetime.strptime(date_str, '%d %b %Y').strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"无法解析Statement Date: {date_str}")
                else:
                    result['extraction_errors'].append("未找到Statement Date")
                
                # 2. Due Date - "Payment Due Date 27 SEP 2024"
                due_match = re.search(r'Payment\s+Due\s+Date\s+(\d{2}\s+[A-Z]{3}\s+\d{4})', text)
                if due_match:
                    due_str = due_match.group(1)
                    try:
                        result['due_date'] = datetime.strptime(due_str, '%d %b %Y').strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"无法解析Due Date: {due_str}")
                else:
                    result['extraction_errors'].append("未找到Payment Due Date")
                
                # 3. Statement Total - "Jumlah Terkini 5,063.02"
                total_match = re.search(r'(?:Total\s+Current\s+Balance|Jumlah\s+Terkini)\s+(\d{1,3}(?:,\d{3})*\.\d{2})', text)
                if total_match:
                    result['statement_total'] = Decimal(total_match.group(1).replace(',', ''))
                else:
                    result['extraction_errors'].append("未找到Statement Total")
                
                # 4. Minimum Payment - "Jumlah Bayaran Minimum 2,956.07"
                min_pay_match = re.search(r'Jumlah\s+Bayaran\s+Minimum\s+(\d{1,3}(?:,\d{3})*\.\d{2})', text)
                if min_pay_match:
                    result['minimum_payment'] = Decimal(min_pay_match.group(1).replace(',', ''))
                else:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class AllianceBankParser(BankParser):
    """Alliance Bank 专用解析器 - 修复版本"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # Statement Date - "Tarikh Penyata 12/08/25" or "Statement Date 12/08/25"
                date_patterns = [
                    r'Tarikh\s+Penyata\s+(\d{2}/\d{2}/\d{2,4})',
                    r'Statement\s+Date\s+(\d{2}/\d{2}/\d{2,4})',
                ]
                for pattern in date_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        date_str = match.group(1)
                        try:
                            # 处理 DD/MM/YY 格式
                            parts = date_str.split('/')
                            if len(parts[2]) == 2:
                                date_str = f"{parts[0]}/{parts[1]}/20{parts[2]}"
                            result['statement_date'] = datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                            break
                        except Exception as e:
                            result['extraction_errors'].append(f"Statement Date解析失败: {date_str} - {e}")
                            continue
                
                # Due Date - "Tarikh Bayaran Perlu Dibuat 01/09/25" or "Payment Due Date 01/09/25"
                due_patterns = [
                    r'Tarikh\s+Bayaran\s+Perlu\s+Dibuat\s+(\d{2}/\d{2}/\d{2,4})',
                    r'Payment\s+Due\s+Date\s+(\d{2}/\d{2}/\d{2,4})',
                ]
                for pattern in due_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        due_str = match.group(1)
                        try:
                            parts = due_str.split('/')
                            if len(parts[2]) == 2:
                                due_str = f"{parts[0]}/{parts[1]}/20{parts[2]}"
                            result['due_date'] = datetime.strptime(due_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                            break
                        except Exception as e:
                            result['extraction_errors'].append(f"Due Date解析失败: {due_str} - {e}")
                            continue
                
                # Current Balance (Statement Total) and Minimum Payment from table row
                # 格式: "CARD_NAME CARD_NUMBER CURRENT_BALANCE MINIMUM_PAYMENT"
                # 例: "YOU:NIQUE MASTERCARD 5465 9464 0768 4514 10,004.46 1,022.72"
                table_match = re.search(r'(\d{4}\s+\d{4}\s+\d{4}\s+\d{4})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})', text)
                if table_match:
                    result['statement_total'] = Decimal(table_match.group(2).replace(',', ''))
                    result['minimum_payment'] = Decimal(table_match.group(3).replace(',', ''))
                
                # 记录缺失字段
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class AmBankParser(BankParser):
    """AmBank / AmBank Islamic 专用解析器"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # AmBank格式需要查看更多页
                for page_num in range(min(3, len(pdf.pages))):
                    page_text = pdf.pages[page_num].extract_text()
                    text += "\n" + page_text
                
                # Statement Date - "Statement Date / Tarikh Penyata 28 MAY 25"
                date_match = re.search(r'Statement\s+Date.*?(\d{2}\s+[A-Z]{3}\s+\d{2,4})', text)
                if date_match:
                    date_str = date_match.group(1)
                    try:
                        # 处理2位或4位年份
                        if len(date_str.split()[-1]) == 2:
                            year = '20' + date_str.split()[-1]
                            date_str = ' '.join(date_str.split()[:-1]) + ' ' + year
                        result['statement_date'] = datetime.strptime(date_str, '%d %b %Y').strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"无法解析Statement Date: {date_str}")
                
                # Due Date - "Payment Due Date / Tarikh Matang Bayaran 17 JUN 25"
                due_match = re.search(r'Payment\s+Due\s+Date.*?(\d{2}\s+[A-Z]{3}\s+\d{2,4})', text)
                if due_match:
                    due_str = due_match.group(1)
                    try:
                        if len(due_str.split()[-1]) == 2:
                            year = '20' + due_str.split()[-1]
                            due_str = ' '.join(due_str.split()[:-1]) + ' ' + year
                        result['due_date'] = datetime.strptime(due_str, '%d %b %Y').strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"无法解析Due Date: {due_str}")
                
                # Statement Total & Minimum Payment - 需要在后续页面找
                # 通常在Summary部分
                total_match = re.search(r'(?:Total\s+Amount\s+Due|New\s+Balance|Outstanding)[:\s]+RM\s*(\d{1,3}(?:,\d{3})*\.?\d{0,2})', text, re.IGNORECASE)
                if total_match:
                    result['statement_total'] = Decimal(total_match.group(1).replace(',', ''))
                
                min_match = re.search(r'(?:Minimum\s+Payment|Bayaran\s+Minimum)[:\s]+RM\s*(\d{1,3}(?:,\d{3})*\.?\d{0,2})', text, re.IGNORECASE)
                if min_match:
                    result['minimum_payment'] = Decimal(min_match.group(1).replace(',', ''))
                
                # 记录缺失字段
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class GenericBankParser(BankParser):
    """通用解析器 - 用于其他银行"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                # 读取前3页
                for page_num in range(min(3, len(pdf.pages))):
                    text += pdf.pages[page_num].extract_text() + "\n"
                
                # 尝试多种日期格式
                date_patterns = [
                    (r'Statement\s+Date[:\s]+(\d{2}\s+[A-Z]{3}\s+\d{2,4})', '%d %b %Y'),
                    (r'Statement\s+Date[:\s]+(\d{2}[/-]\d{2}[/-]\d{4})', '%d-%m-%Y'),
                    (r'Tarikh\s+Penyata[:\s]+(\d{2}[/-]\d{2}[/-]\d{4})', '%d-%m-%Y'),
                ]
                
                for pattern, date_format in date_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        try:
                            date_str = match.group(1).replace('/', '-')
                            if len(date_str.split()[-1]) == 2:  # 处理2位年份
                                year = '20' + date_str.split()[-1]
                                date_str = ' '.join(date_str.split()[:-1]) + ' ' + year
                            result['statement_date'] = datetime.strptime(date_str, date_format).strftime('%Y-%m-%d')
                            break
                        except:
                            continue
                
                # Due Date
                due_patterns = [
                    (r'Payment\s+Due\s+Date[:\s]+(\d{2}\s+[A-Z]{3}\s+\d{2,4})', '%d %b %Y'),
                    (r'Payment\s+Due\s+Date[:\s]+(\d{2}[/-]\d{2}[/-]\d{4})', '%d-%m-%Y'),
                    (r'Due\s+Date[:\s]+(\d{2}[/-]\d{2}[/-]\d{4})', '%d-%m-%Y'),
                ]
                
                for pattern, date_format in due_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        try:
                            due_str = match.group(1).replace('/', '-')
                            if len(due_str.split()[-1]) == 2:
                                year = '20' + due_str.split()[-1]
                                due_str = ' '.join(due_str.split()[:-1]) + ' ' + year
                            result['due_date'] = datetime.strptime(due_str, date_format).strftime('%Y-%m-%d')
                            break
                        except:
                            continue
                
                # Statement Total
                total_patterns = [
                    r'(?:Total\s+Amount\s+Due|New\s+Balance|Outstanding\s+Balance|Current\s+Balance)[:\s]+RM\s*(\d{1,3}(?:,\d{3})*\.?\d{0,2})',
                    r'(?:Jumlah.*?Terkini)[:\s]+(\d{1,3}(?:,\d{3})*\.?\d{0,2})',
                ]
                
                for pattern in total_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        result['statement_total'] = Decimal(match.group(1).replace(',', ''))
                        break
                
                # Minimum Payment
                min_patterns = [
                    r'(?:Minimum\s+Payment|Minimum\s+Amount\s+Due)[:\s]+RM\s*(\d{1,3}(?:,\d{3})*\.?\d{0,2})',
                    r'(?:Bayaran\s+Minimum|Jumlah\s+Bayaran\s+Minimum)[:\s]+(\d{1,3}(?:,\d{3})*\.?\d{0,2})',
                ]
                
                for pattern in min_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        result['minimum_payment'] = Decimal(match.group(1).replace(',', ''))
                        break
                
                # 记录缺失字段
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class OCBCBankParser(BankParser):
    """OCBC Bank 专用解析器"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # OCBC格式: Statement Date: 13 MAY 2025
                date_match = re.search(r'Statement Date[:\s]+(\d{1,2}\s+[A-Z]{3}\s+\d{4})', text, re.IGNORECASE)
                if date_match:
                    date_str = date_match.group(1)
                    dt = datetime.strptime(date_str, '%d %b %Y')
                    result['statement_date'] = dt.strftime('%Y-%m-%d')
                
                # Due Date: Payment Due Date 02 JUN 2025
                due_match = re.search(r'Payment Due Date[:\s]+(\d{1,2}\s+[A-Z]{3}\s+\d{4})', text, re.IGNORECASE)
                if due_match:
                    date_str = due_match.group(1)
                    dt = datetime.strptime(date_str, '%d %b %Y')
                    result['due_date'] = dt.strftime('%Y-%m-%d')
                
                # Current Balance and Minimum Payment from table row
                # 格式: CARD_NAME CARD_NUMBER BALANCE MIN_PAYMENT
                # 例: GE MASTERCARD PLATINUM 5401-6200-0093-3506 1,190.69 446.00
                table_match = re.search(r'(\d{4}-\d{4}-\d{4}-\d{4})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})', text)
                if table_match:
                    result['statement_total'] = table_match.group(2).replace(',', '')
                    result['minimum_payment'] = table_match.group(3).replace(',', '')
                
                # 记录缺失字段
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class UOBBankParser(BankParser):
    """UOB Bank 专用解析器 - 从真实PDF提取数据（无估算）"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                page1_text = pdf.pages[0].extract_text()
                
                # Page 1: Statement Date (支持短格式 "13 MAY 25" 和长格式 "13 May 2025")
                date_match = re.search(r'Statement Date[:\s]+(\d{1,2}\s+[A-Z]{3}\s+\d{2,4})', page1_text, re.IGNORECASE)
                if date_match:
                    date_str = date_match.group(1)
                    try:
                        if len(date_str.split()[-1]) == 2:
                            dt = datetime.strptime(date_str, '%d %b %y')
                        else:
                            dt = datetime.strptime(date_str, '%d %b %Y')
                        result['statement_date'] = dt.strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"Statement Date格式错误: {date_str}")
                
                # Page 1: Payment Due Date (支持短格式 "02 JUN 25" 和长格式 "02 June 2025")
                due_match = re.search(r'Payment Due Date[:\s]+(\d{1,2}\s+[A-Z]{3}\s+\d{2,4})', page1_text, re.IGNORECASE)
                if due_match:
                    date_str = due_match.group(1)
                    try:
                        if len(date_str.split()[-1]) == 2:
                            dt = datetime.strptime(date_str, '%d %b %y')
                        else:
                            dt = datetime.strptime(date_str, '%d %b %Y')
                        result['due_date'] = dt.strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"Due Date格式错误: {date_str}")
                
                # 遍历所有页面查找 Minimum Payment Due 和 Total Balance Due（真实打印值）
                for page in pdf.pages:
                    text = page.extract_text()
                    lines = text.split('\n')
                    
                    # 查找 "Minimum Payment Due 10.38" 格式（PDF真实打印的值）
                    if result['minimum_payment'] is None:
                        min_match = re.search(r'Minimum Payment Due[:\s]+([\d,]+\.\d{2})', text, re.IGNORECASE)
                        if min_match:
                            result['minimum_payment'] = min_match.group(1).replace(',', '')
                    
                    # 查找 Total Balance Due（跨行格式：标题在一行，数值在后续行）
                    if result['statement_total'] is None:
                        for i, line in enumerate(lines):
                            if 'Total Balance Due' in line and i + 3 < len(lines):
                                # 查找标题行后3行内的数字行（通常是表格数据行）
                                for j in range(i+1, min(i+4, len(lines))):
                                    # 匹配包含多个数字的行，取最后一个数字（Total Balance Due列）
                                    number_match = re.findall(r'([\d,]+\.\d{2})', lines[j])
                                    if number_match:
                                        result['statement_total'] = number_match[-1].replace(',', '')
                                        break
                                if result['statement_total']:
                                    break
                
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment Due（已检查所有页面）")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class HSBCBankParser(BankParser):
    """HSBC Bank 专用解析器 - 从真实PDF提取数据（无估算）"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # Statement Date 和 Due Date 在同一行： "Statement Date 13 May 2025 Payment Due Date 03 June 2025"
                # 或缩写格式："Statement Date 13 Oct 2025 Payment Due Date 02 November 2025"
                date_line_match = re.search(r'Statement Date[:\s]+(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})\s+Payment Due Date[:\s]+(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})', text, re.IGNORECASE)
                if date_line_match:
                    try:
                        stmt_date_str = date_line_match.group(1)
                        # 尝试两种格式：完整月份名(%B: May, June)和缩写月份名(%b: Oct, Nov)
                        try:
                            dt = datetime.strptime(stmt_date_str, '%d %B %Y')
                        except:
                            dt = datetime.strptime(stmt_date_str, '%d %b %Y')
                        result['statement_date'] = dt.strftime('%Y-%m-%d')
                        
                        due_date_str = date_line_match.group(2)
                        try:
                            dt = datetime.strptime(due_date_str, '%d %B %Y')
                        except:
                            dt = datetime.strptime(due_date_str, '%d %b %Y')
                        result['due_date'] = dt.strftime('%Y-%m-%d')
                    except Exception as e:
                        result['extraction_errors'].append(f"日期格式错误: {date_line_match.group(0)} - {str(e)}")
                
                # 如果上面的combined pattern没找到，尝试单独查找
                if result['statement_date'] is None:
                    date_match = re.search(r'Statement Date[:\s]+(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})', text, re.IGNORECASE)
                    if date_match:
                        try:
                            date_str = date_match.group(1)
                            try:
                                dt = datetime.strptime(date_str, '%d %B %Y')
                            except:
                                dt = datetime.strptime(date_str, '%d %b %Y')
                            result['statement_date'] = dt.strftime('%Y-%m-%d')
                        except Exception as e:
                            result['extraction_errors'].append(f"Statement Date格式错误: {date_match.group(1)} - {str(e)}")
                
                if result['due_date'] is None:
                    due_match = re.search(r'Payment Due Date[:\s]+(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})', text, re.IGNORECASE)
                    if due_match:
                        try:
                            date_str = due_match.group(1)
                            dt = datetime.strptime(date_str, '%d %B %Y')
                            result['due_date'] = dt.strftime('%Y-%m-%d')
                        except:
                            result['extraction_errors'].append(f"Due Date格式错误: {due_match.group(1)}")
                
                # 查找表格行： 卡号 余额 最低还款 超限额 应付金额
                # 例: 4364800001380034 50.00 50.00 0.00 50.00
                # 格式：16位卡号 + 4个数字（Balance, Minimum Payment, Overlimit, Payment Due）
                table_match = re.search(r'(\d{16})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})', text)
                if table_match:
                    result['statement_total'] = table_match.group(2).replace(',', '')  # Balance
                    result['minimum_payment'] = table_match.group(3).replace(',', '')  # Minimum Monthly Payment
                else:
                    # 回退方案：尝试传统的单独匹配
                    if result['statement_total'] is None:
                        total_match = re.search(r'(?:Balance|New\s+Balance)[:\s]+(?:RM\s*|MYR\s*)?([\d,]+\.\d{2})', text, re.IGNORECASE)
                        if total_match:
                            result['statement_total'] = total_match.group(1).replace(',', '')
                    
                    if result['minimum_payment'] is None:
                        min_match = re.search(r'Minimum\s+(?:Monthly\s+)?Payment[:\s]+(?:RM\s*|MYR\s*)?([\d,]+\.\d{2})', text, re.IGNORECASE)
                        if min_match:
                            result['minimum_payment'] = min_match.group(1).replace(',', '')
                
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                if result['due_date'] is None:
                    result['extraction_errors'].append("未找到Due Date")
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class StandardCharteredBankParser(BankParser):
    """Standard Chartered Bank 专用解析器 - 从真实PDF提取数据（无估算）"""
    
    def parse(self, pdf_path: str) -> Dict:
        result = {
            'statement_date': None,
            'due_date': None,
            'statement_total': None,
            'minimum_payment': None,
            'extraction_errors': []
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                lines = text.split('\n')
                
                # Statement Date: "Statement Date / Tarikh Penyata : 14 Oct 2025"
                stmt_date_match = re.search(r'Statement Date[:/\s]+.*?:\s*(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})', text, re.IGNORECASE)
                if stmt_date_match:
                    try:
                        date_str = stmt_date_match.group(1)
                        dt = datetime.strptime(date_str, '%d %b %Y')
                        result['statement_date'] = dt.strftime('%Y-%m-%d')
                    except:
                        result['extraction_errors'].append(f"Statement Date格式错误: {stmt_date_match.group(1)}")
                
                # Due Date: "Payment Due Date / Tarikh Akhir : IMMEDIATE" 或 "Payment Due Date / Tarikh Akhir : 15 Nov 2025"
                # 注意：IMMEDIATE不是日期，将其留空（NULL）
                due_date_match = re.search(r'Payment Due Date[:/\s]+.*?:\s*(.+)', text, re.IGNORECASE)
                if due_date_match:
                    due_value = due_date_match.group(1).strip()
                    if 'IMMEDIATE' not in due_value.upper():
                        # 尝试解析为日期
                        try:
                            date_match = re.search(r'(\d{1,2}\s+[A-Z][a-z]{2,8}\s+\d{4})', due_value, re.IGNORECASE)
                            if date_match:
                                dt = datetime.strptime(date_match.group(1), '%d %b %Y')
                                result['due_date'] = dt.strftime('%Y-%m-%d')
                        except:
                            result['extraction_errors'].append(f"Due Date格式错误: {due_value}")
                    # 如果是IMMEDIATE，due_date保持None
                
                # Minimum Payment和Balance：查找卡片名称行后的数字
                # 格式: "SIMPLY CASH CREDIT CARD 73,889.93 9,220.36"
                # 或: "CARD_NAME NEW_BALANCE MIN_PAYMENT"
                for i, line in enumerate(lines):
                    if any(card_type in line.upper() for card_type in ['CREDIT CARD', 'VISA', 'MASTERCARD', 'AMEX']):
                        # 提取该行的所有数字
                        numbers = re.findall(r'([\d,]+\.\d{2})', line)
                        if len(numbers) >= 2:
                            result['statement_total'] = numbers[-2].replace(',', '')  # 倒数第二个数字是balance
                            result['minimum_payment'] = numbers[-1].replace(',', '')  # 最后一个数字是minimum payment
                            break
                
                # 验证必填字段
                if result['statement_date'] is None:
                    result['extraction_errors'].append("未找到Statement Date")
                # due_date可以为空（IMMEDIATE情况）
                if result['statement_total'] is None:
                    result['extraction_errors'].append("未找到Statement Total")
                if result['minimum_payment'] is None:
                    result['extraction_errors'].append("未找到Minimum Payment")
                    
        except Exception as e:
            result['extraction_errors'].append(f"PDF读取失败: {str(e)}")
        
        return result

class PDFFieldExtractor:
    """PDF字段提取协调器"""
    
    def __init__(self):
        # 银行解析器映射
        self.parsers = {
            'Hong Leong Bank': HongLeongBankParser('Hong Leong Bank'),
            'HLB': HongLeongBankParser('HLB'),
            'HONG_LEONG': HongLeongBankParser('HONG_LEONG'),
            'Alliance Bank': AllianceBankParser('Alliance Bank'),
            'Alliance': AllianceBankParser('Alliance'),
            'AmBank': AmBankParser('AmBank'),
            'AMBANK': AmBankParser('AMBANK'),
            'AmBank Islamic': AmBankParser('AmBank Islamic'),
            'OCBC': OCBCBankParser('OCBC'),
            'UOB': UOBBankParser('UOB'),
            'HSBC': HSBCBankParser('HSBC'),
            'STANDARD CHARTERED': StandardCharteredBankParser('STANDARD CHARTERED'),
            'Standard Chartered': StandardCharteredBankParser('Standard Chartered'),
            'SC': StandardCharteredBankParser('SC'),
        }
    
    def get_parser(self, bank_name: str) -> BankParser:
        """获取银行对应的解析器"""
        if bank_name in self.parsers:
            return self.parsers[bank_name]
        # 默认使用通用解析器
        return GenericBankParser(bank_name)
    
    def extract_fields(self, pdf_path: str, bank_name: str) -> Dict:
        """提取PDF字段"""
        parser = self.get_parser(bank_name)
        return parser.parse(pdf_path)

def generate_comparison_report(output_file: str = 'pdf_vs_database_comparison.xlsx'):
    """生成PDF vs 数据库对比报告"""
    
    conn = sqlite3.connect('db/smart_loan_manager.db')
    cursor = conn.cursor()
    
    # 获取所有PDF记录
    cursor.execute("""
        SELECT 
            s.id,
            cu.name as customer_name,
            c.bank_name,
            c.card_number_last4,
            s.statement_date as db_statement_date,
            s.due_date as db_due_date,
            s.statement_total as db_statement_total,
            s.minimum_payment as db_minimum_payment,
            s.file_path
        FROM statements s
        JOIN credit_cards c ON s.card_id = c.id
        JOIN customers cu ON c.customer_id = cu.id
        WHERE s.file_path IS NOT NULL 
        AND s.file_path LIKE '%.pdf'
        ORDER BY cu.name, c.bank_name, s.statement_date
    """)
    
    records = cursor.fetchall()
    conn.close()
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF vs Database对比"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 50
    
    # 标题行
    headers = [
        'ID', 'Customer', 'Bank', 'Card',
        'DB Stmt Date', 'PDF Stmt Date',
        'DB Due Date', 'PDF Due Date',
        'DB Total', 'PDF Total',
        'DB Min Payment', 'PDF Min Payment',
        'Status', 'Errors'
    ]
    
    header_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    header_font = Font(bold=True, color='000000', size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 初始化提取器
    extractor = PDFFieldExtractor()
    
    # 处理每条记录
    row_num = 2
    total_records = len(records)
    correct_count = 0
    error_count = 0
    
    print(f"\n🔍 开始解析 {total_records} 个PDF文件...\n")
    
    for idx, record in enumerate(records, 1):
        stmt_id, customer, bank, card, db_stmt_date, db_due_date, db_total, db_min_pay, pdf_path = record
        
        print(f"[{idx}/{total_records}] {customer} - {bank} - {card}")
        
        # 检查文件是否存在
        if not os.path.exists(pdf_path):
            ws.cell(row=row_num, column=1, value=stmt_id)
            ws.cell(row=row_num, column=2, value=customer)
            ws.cell(row=row_num, column=3, value=bank)
            ws.cell(row=row_num, column=4, value=card)
            ws.cell(row=row_num, column=13, value='文件不存在')
            ws.cell(row=row_num, column=14, value=pdf_path)
            
            # 标红
            for col in range(1, 15):
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color='FFcccc', end_color='FFcccc', fill_type='solid')
            
            row_num += 1
            error_count += 1
            continue
        
        # 提取PDF字段
        pdf_data = extractor.extract_fields(pdf_path, bank)
        
        # 填充Excel
        ws.cell(row=row_num, column=1, value=stmt_id)
        ws.cell(row=row_num, column=2, value=customer)
        ws.cell(row=row_num, column=3, value=bank)
        ws.cell(row=row_num, column=4, value=card)
        
        ws.cell(row=row_num, column=5, value=db_stmt_date)
        ws.cell(row=row_num, column=6, value=pdf_data['statement_date'])
        
        ws.cell(row=row_num, column=7, value=db_due_date)
        ws.cell(row=row_num, column=8, value=pdf_data['due_date'])
        
        ws.cell(row=row_num, column=9, value=float(db_total) if db_total else None)
        ws.cell(row=row_num, column=10, value=float(pdf_data['statement_total']) if pdf_data['statement_total'] else None)
        
        ws.cell(row=row_num, column=11, value=float(db_min_pay) if db_min_pay else None)
        ws.cell(row=row_num, column=12, value=float(pdf_data['minimum_payment']) if pdf_data['minimum_payment'] else None)
        
        # 判断状态
        has_errors = len(pdf_data['extraction_errors']) > 0
        has_differences = False
        
        # 比较字段
        if pdf_data['statement_date'] and db_stmt_date != pdf_data['statement_date']:
            has_differences = True
        if pdf_data['due_date'] and db_due_date != pdf_data['due_date']:
            has_differences = True
        if pdf_data['statement_total'] and db_total:
            if abs(float(db_total) - float(pdf_data['statement_total'])) > 0.01:
                has_differences = True
        if pdf_data['minimum_payment'] and db_min_pay:
            if abs(float(db_min_pay) - float(pdf_data['minimum_payment'])) > 0.01:
                has_differences = True
        
        if has_errors:
            status = '❌ 提取失败'
            fill_color = 'FFcccc'  # 红色
            error_count += 1
        elif has_differences:
            status = '⚠️ 有差异'
            fill_color = 'FFffcc'  # 黄色
        else:
            status = '✅ 正确'
            fill_color = 'ccFFcc'  # 绿色
            correct_count += 1
        
        ws.cell(row=row_num, column=13, value=status)
        ws.cell(row=row_num, column=14, value='; '.join(pdf_data['extraction_errors']))
        
        # 设置颜色
        for col in range(1, 15):
            ws.cell(row=row_num, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        row_num += 1
    
    # 保存Excel
    wb.save(output_file)
    
    print(f"\n{'='*100}")
    print(f"📊 对比报告生成完成")
    print(f"{'='*100}\n")
    print(f"  ✅ 完全正确: {correct_count} 条")
    print(f"  ⚠️  有差异: {total_records - correct_count - error_count} 条")
    print(f"  ❌ 提取失败: {error_count} 条")
    print(f"  📄 总计: {total_records} 条")
    print(f"\n📁 报告文件: {output_file}\n")
    print(f"{'='*100}\n")

if __name__ == "__main__":
    generate_comparison_report()
