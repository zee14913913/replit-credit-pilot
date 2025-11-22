#!/usr/bin/env python3
"""
专业会计文件组织系统
===================================
为每个客户创建独立的会计文件夹结构
符合专业会计审计标准和客户保密要求

文件夹结构:
/accounting_files/[CUSTOMER_NAME]/
    /monthly_statements/          # 月结单Excel文件
        2024-09_Alliance_Bank_Statement.xlsx
        2024-09_HSBC_Statement.xlsx
        2024-09_Summary.xlsx
    /transaction_details/         # 交易明细
        2024-09_All_Transactions.xlsx
    /transfer_records/           # 转账记录
        Transfer_Log.xlsx
    /reports/                    # 汇总报告
        Annual_Summary_2024.xlsx
        GZ_Settlement_Report.xlsx
"""

import sqlite3
import os
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
from decimal import Decimal

class CustomerAccountingOrganizer:
    """客户会计文件组织器"""
    
    def __init__(self, db_path='db/smart_loan_manager.db'):
        self.db_path = db_path
        self.base_dir = Path('accounting_files')
        self.conn = None
        
    def connect_db(self):
        """连接数据库"""
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        
    def close_db(self):
        """关闭数据库"""
        if self.conn:
            self.conn.close()
    
    def get_all_customers(self):
        """获取所有客户"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT id, name, customer_code 
            FROM customers 
            WHERE customer_code IS NOT NULL
        """)
        return cursor.fetchall()
    
    def create_customer_folder_structure(self, customer_name):
        """创建客户文件夹结构"""
        customer_dir = self.base_dir / customer_name
        
        # 创建子文件夹
        folders = [
            customer_dir / 'monthly_statements',
            customer_dir / 'transaction_details',
            customer_dir / 'transfer_records',
            customer_dir / 'reports',
            customer_dir / 'source_pdfs'
        ]
        
        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)
        
        print(f"  ✅ 创建文件夹结构: {customer_dir}")
        return customer_dir
    
    def get_customer_monthly_statements(self, customer_id):
        """获取客户的月结单数据"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT 
                bank_name,
                statement_month,
                previous_balance_total,
                closing_balance_total,
                owner_balance,
                gz_balance,
                owner_expenses,
                owner_payments,
                gz_expenses,
                gz_payments,
                transaction_count,
                file_paths
            FROM monthly_statements
            WHERE customer_id = ?
            ORDER BY statement_month, bank_name
        """, (customer_id,))
        return cursor.fetchall()
    
    def get_customer_transactions(self, customer_id, statement_month=None):
        """获取客户的交易明细"""
        cursor = self.conn.cursor()
        
        if statement_month:
            cursor.execute("""
                SELECT 
                    t.transaction_date,
                    t.description,
                    CASE WHEN t.transaction_type = 'purchase' THEN t.amount ELSE 0 END as dr,
                    CASE WHEN t.transaction_type = 'payment' THEN t.amount ELSE 0 END as cr,
                    t.category,
                    ms.bank_name,
                    t.owner_flag
                FROM transactions t
                JOIN monthly_statements ms ON t.monthly_statement_id = ms.id
                WHERE ms.customer_id = ? AND ms.statement_month = ?
                ORDER BY ms.bank_name, t.transaction_date
            """, (customer_id, statement_month))
        else:
            cursor.execute("""
                SELECT 
                    t.transaction_date,
                    t.description,
                    CASE WHEN t.transaction_type = 'purchase' THEN t.amount ELSE 0 END as dr,
                    CASE WHEN t.transaction_type = 'payment' THEN t.amount ELSE 0 END as cr,
                    t.category,
                    ms.bank_name,
                    ms.statement_month,
                    t.owner_flag
                FROM transactions t
                JOIN monthly_statements ms ON t.monthly_statement_id = ms.id
                WHERE ms.customer_id = ?
                ORDER BY ms.statement_month, ms.bank_name, t.transaction_date
            """, (customer_id,))
        
        return cursor.fetchall()
    
    def create_monthly_statement_excel(self, customer_dir, bank_name, statement_month, stmt_data):
        """创建单个银行的月结单Excel"""
        filename = f"{statement_month}_{bank_name.replace(' ', '_')}_Statement.xlsx"
        filepath = customer_dir / 'monthly_statements' / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "月结单"
        
        # 样式定义
        header_fill = PatternFill(start_color="FF007F", end_color="FF007F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(size=14, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws['A1'] = "信用卡月结单"
        ws['A1'].font = title_font
        ws['A2'] = f"银行: {bank_name}"
        ws['A3'] = f"月份: {statement_month}"
        ws['A4'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 余额信息
        row = 6
        headers = ['项目', '金额 (RM)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 余额数据
        balance_data = [
            ('期初余额', stmt_data['previous_balance_total'] or 0),
            ('期末余额', stmt_data['closing_balance_total'] or 0),
            ('OWNER 余额', stmt_data['owner_balance'] or 0),
            ('GZ 余额', stmt_data['gz_balance'] or 0),
            ('OWNER 费用', stmt_data['owner_expenses'] or 0),
            ('OWNER 付款', stmt_data['owner_payments'] or 0),
            ('GZ 费用', stmt_data['gz_expenses'] or 0),
            ('GZ 付款', stmt_data['gz_payments'] or 0),
            ('GZ 1% 管理费', (stmt_data['gz_expenses'] or 0) * 0.01),
            ('交易笔数', stmt_data['transaction_count'] or 0)
        ]
        
        for idx, (label, value) in enumerate(balance_data, 1):
            ws.cell(row=row+idx, column=1, value=label).border = border
            cell_value = ws.cell(row=row+idx, column=2, value=value)
            cell_value.border = border
            if isinstance(value, (int, float)) and label != '交易笔数':
                cell_value.number_format = '#,##0.00'
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        wb.save(filepath)
        return filepath
    
    def create_monthly_summary_excel(self, customer_dir, customer_name, statement_month, monthly_data):
        """创建月度汇总Excel（所有银行）"""
        filename = f"{statement_month}_Summary.xlsx"
        filepath = customer_dir / 'monthly_statements' / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "月度汇总"
        
        # 样式
        header_fill = PatternFill(start_color="322446", end_color="322446", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        owner_fill = PatternFill(start_color="FFE0F0", end_color="FFE0F0", fill_type="solid")
        gz_fill = PatternFill(start_color="E8D8F0", end_color="E8D8F0", fill_type="solid")
        
        # 标题
        ws['A1'] = f"{customer_name} - {statement_month} 月度汇总"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # 表头
        headers = ['银行', 'OWNER费用', 'OWNER付款', 'GZ费用', 'GZ付款', 'GZ 1%费', '交易数', '总余额']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        row = 4
        total_owner_exp = 0
        total_owner_pay = 0
        total_gz_exp = 0
        total_gz_pay = 0
        total_txn = 0
        
        for stmt in monthly_data:
            ws.cell(row=row, column=1, value=stmt['bank_name'])
            ws.cell(row=row, column=2, value=stmt['owner_expenses'] or 0).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=stmt['owner_payments'] or 0).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=stmt['gz_expenses'] or 0).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=stmt['gz_payments'] or 0).number_format = '#,##0.00'
            
            gz_fee = (stmt['gz_expenses'] or 0) * 0.01
            ws.cell(row=row, column=6, value=gz_fee).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=stmt['transaction_count'] or 0)
            ws.cell(row=row, column=8, value=stmt['closing_balance_total'] or 0).number_format = '#,##0.00'
            
            # Owner行高亮
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = owner_fill
            
            total_owner_exp += (stmt['owner_expenses'] or 0)
            total_owner_pay += (stmt['owner_payments'] or 0)
            total_gz_exp += (stmt['gz_expenses'] or 0)
            total_gz_pay += (stmt['gz_payments'] or 0)
            total_txn += (stmt['transaction_count'] or 0)
            
            row += 1
        
        # 合计行
        ws.cell(row=row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_owner_exp).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=total_owner_pay).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=total_gz_exp).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=total_gz_pay).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=total_gz_exp * 0.01).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=total_txn)
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = gz_fill
        
        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[chr(64+col)].width = 15
        
        wb.save(filepath)
        return filepath
    
    def create_transaction_details_excel(self, customer_dir, customer_name, statement_month, transactions):
        """创建交易明细Excel"""
        filename = f"{statement_month}_All_Transactions.xlsx"
        filepath = customer_dir / 'transaction_details' / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "交易明细"
        
        # 标题
        ws['A1'] = f"{customer_name} - {statement_month} 完整交易明细"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # 表头
        headers = ['日期', '银行', '描述', '借记(DR)', '贷记(CR)', '分类', '归属']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = PatternFill(start_color="FF007F", end_color="FF007F", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # 数据
        row = 4
        for txn in transactions:
            ws.cell(row=row, column=1, value=txn[0])  # transaction_date
            ws.cell(row=row, column=2, value=txn[5])  # bank_name
            ws.cell(row=row, column=3, value=txn[1])  # description
            ws.cell(row=row, column=4, value=txn[2] or 0).number_format = '#,##0.00'  # dr
            ws.cell(row=row, column=5, value=txn[3] or 0).number_format = '#,##0.00'  # cr
            ws.cell(row=row, column=6, value=txn[4] or 'Other')  # category
            
            # 使用owner_flag
            owner = txn[6] if len(txn) > 6 and txn[6] else 'OWNER'
            if owner == 'INFINITE':
                owner = 'GZ'
            
            ws.cell(row=row, column=7, value=owner)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        
        wb.save(filepath)
        return filepath
    
    def process_customer(self, customer_id, customer_name, customer_code):
        """处理单个客户的所有数据"""
        print(f"\n{'='*80}")
        print(f"📁 处理客户: {customer_name} ({customer_code})")
        print(f"{'='*80}")
        
        # 创建文件夹结构
        customer_dir = self.create_customer_folder_structure(customer_name)
        
        # 获取月结单数据
        statements = self.get_customer_monthly_statements(customer_id)
        
        if not statements:
            print(f"  ⚠️ 没有找到月结单数据")
            return
        
        print(f"  📊 找到 {len(statements)} 条月结单记录")
        
        # 按月份分组
        monthly_groups = defaultdict(list)
        for stmt in statements:
            monthly_groups[stmt['statement_month']].append(stmt)
        
        created_files = []
        
        # 为每个月份创建文件
        for month, month_statements in sorted(monthly_groups.items()):
            print(f"\n  📅 处理 {month}:")
            
            # 创建每个银行的独立月结单
            for stmt in month_statements:
                filepath = self.create_monthly_statement_excel(
                    customer_dir,
                    stmt['bank_name'],
                    month,
                    stmt
                )
                created_files.append(filepath)
                print(f"    ✅ {filepath.name}")
            
            # 创建月度汇总
            summary_path = self.create_monthly_summary_excel(
                customer_dir,
                customer_name,
                month,
                month_statements
            )
            created_files.append(summary_path)
            print(f"    ✅ {summary_path.name}")
            
            # 创建交易明细
            transactions = self.get_customer_transactions(customer_id, month)
            if transactions:
                details_path = self.create_transaction_details_excel(
                    customer_dir,
                    customer_name,
                    month,
                    transactions
                )
                created_files.append(details_path)
                print(f"    ✅ {details_path.name} ({len(transactions)} 笔交易)")
        
        print(f"\n  ✅ 共创建 {len(created_files)} 个文件")
        return customer_dir, created_files
    
    def run(self):
        """运行完整的文件组织流程"""
        print("\n🚀 开始创建专业会计文件组织系统")
        print("="*80)
        
        self.connect_db()
        
        try:
            # 获取所有客户
            customers = self.get_all_customers()
            print(f"找到 {len(customers)} 个客户\n")
            
            all_results = {}
            
            for customer in customers:
                result = self.process_customer(
                    customer['id'],
                    customer['name'],
                    customer['customer_code']
                )
                if result:
                    all_results[customer['name']] = result
            
            # 总结
            print(f"\n{'='*80}")
            print("📊 文件组织完成总结")
            print(f"{'='*80}")
            for customer_name, (customer_dir, files) in all_results.items():
                print(f"\n✅ {customer_name}")
                print(f"   📁 目录: {customer_dir}")
                print(f"   📄 文件数: {len(files)}")
            
            print(f"\n{'='*80}")
            print("✅ 专业会计文件组织系统创建完成！")
            print(f"{'='*80}\n")
            
        finally:
            self.close_db()


def main():
    """主函数"""
    organizer = CustomerAccountingOrganizer()
    organizer.run()


if __name__ == '__main__':
    main()
