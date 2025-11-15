"""
Chang Choon Chow 详细分类表格报告生成器（Excel格式）
按银行、月份生成详细的消费、付款、转账、汇总表
"""
import sqlite3
from decimal import Decimal
from collections import defaultdict
from datetime import datetime
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl未安装，正在安装...")
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl'], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Supplier List（7个供应商）
SUPPLIER_LIST = ['7SL', 'DINAS', 'RAUB SYC HAINAN', 'AI SMART TECH', 'HUAWEI', 'PASAR RAYA', 'PUCHONG HERBS']

# GZ关键词（付款识别）
GZ_PAYMENT_KEYWORDS = ['GZ', 'KENG CHOW', 'INFINITE']

def is_supplier(description):
    """检查是否为Supplier消费"""
    desc_upper = description.upper()
    for supplier in SUPPLIER_LIST:
        if supplier.upper() in desc_upper:
            return True, supplier
    return False, None

def is_gz_payment(description):
    """检查是否为GZ付款"""
    desc_upper = description.upper()
    for keyword in GZ_PAYMENT_KEYWORDS:
        if keyword in desc_upper:
            return True
    return False

def classify_transaction(description, amount):
    """分类交易"""
    desc_upper = description.upper()
    
    # 1. 检查是否为付款
    if 'PAYMENT' in desc_upper or 'THANK YOU' in desc_upper:
        if is_gz_payment(description):
            return 'GZ_PAYMENT', 'DIRECT', {}
        else:
            return 'OWNER_PAYMENT', None, {}
    
    # 2. 检查是否为Supplier消费
    is_sup, supplier_name = is_supplier(description)
    if is_sup:
        fee = amount * Decimal('0.01')
        return 'GZ_EXPENSE', 'SUPPLIER', {'supplier': supplier_name, 'fee': fee}
    
    # 3. 其他消费归为Owner
    return 'OWNER_EXPENSE', None, {}

def create_styled_header(ws, headers, row=1, bg_color='322446'):
    """创建样式化的表头"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_excel_report():
    """生成Excel格式详细报告"""
    conn = sqlite3.connect('db/smart_loan_manager.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 查询所有Chang Choon Chow的交易
    cursor.execute("""
        SELECT 
            t.transaction_date,
            t.description,
            t.amount,
            t.category,
            m.bank_name,
            m.statement_month,
            m.id as monthly_statement_id
        FROM transactions t
        JOIN monthly_statements m ON t.monthly_statement_id = m.id
        WHERE m.customer_id = 10
        ORDER BY m.statement_month, m.bank_name, t.transaction_date
    """)
    
    transactions = cursor.fetchall()
    
    # 创建Excel工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 按银行和月份分组
    grouped = defaultdict(lambda: defaultdict(list))
    for txn in transactions:
        bank = txn['bank_name']
        month = txn['statement_month']
        grouped[bank][month].append(txn)
    
    # 创建汇总表
    summary_ws = wb.create_sheet('汇总报告', 0)
    summary_row = 1
    
    # 标题
    summary_ws.merge_cells(f'A{summary_row}:G{summary_row}')
    title_cell = summary_ws.cell(row=summary_row, column=1, value='Chang Choon Chow 结算报告汇总')
    title_cell.font = Font(bold=True, size=16, color='FF007F')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_row += 2
    
    # 客户信息
    info_data = [
        ['客户代码:', 'Be_rich_CCC'],
        ['客户姓名:', 'Chang Choon Chow'],
        ['报告时间:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['交易总数:', f'{len(transactions)}笔'],
        ['银行数量:', f'{len(grouped)}家'],
    ]
    
    for info in info_data:
        summary_ws.cell(row=summary_row, column=1, value=info[0]).font = Font(bold=True)
        summary_ws.cell(row=summary_row, column=2, value=info[1])
        summary_row += 1
    
    summary_row += 2
    
    # 按银行月份汇总表头
    create_styled_header(summary_ws, 
                        ['银行', '月份', 'Owner消费', 'GZ Supplier消费', '1% Fee', 'Owner付款', 'GZ付款'],
                        summary_row)
    summary_row += 1
    
    # 总统计
    grand_total = {
        'owner_expenses': Decimal('0'),
        'gz_expenses': Decimal('0'),
        'fees': Decimal('0'),
        'owner_payments': Decimal('0'),
        'gz_payments': Decimal('0')
    }
    
    # 按银行生成详细表
    banks = sorted(grouped.keys())
    
    for bank in banks:
        months = sorted(grouped[bank].keys())
        
        for month in months:
            txns = grouped[bank][month]
            
            # 创建银行月份工作表
            sheet_name = f"{bank[:10]}_{month}"
            ws = wb.create_sheet(sheet_name)
            
            row = 1
            
            # 标题
            ws.merge_cells(f'A{row}:F{row}')
            title = ws.cell(row=row, column=1, value=f'{bank} - {month}')
            title.font = Font(bold=True, size=14, color='FF007F')
            title.alignment = Alignment(horizontal='center')
            row += 2
            
            # 分类交易
            owner_expenses = []
            gz_expenses = []
            owner_payments = []
            gz_payments = []
            
            month_stats = {
                'owner_expenses': Decimal('0'),
                'gz_expenses': Decimal('0'),
                'fees': Decimal('0'),
                'owner_payments': Decimal('0'),
                'gz_payments': Decimal('0')
            }
            
            for txn in txns:
                amount = abs(Decimal(str(txn['amount'])))
                txn_type, subtype, details = classify_transaction(txn['description'], amount)
                
                txn_data = {
                    'date': txn['transaction_date'],
                    'description': txn['description'],
                    'amount': amount
                }
                
                if txn_type == 'OWNER_EXPENSE':
                    owner_expenses.append(txn_data)
                    month_stats['owner_expenses'] += amount
                    
                elif txn_type == 'GZ_EXPENSE':
                    txn_data['supplier'] = details.get('supplier', 'Unknown')
                    txn_data['fee'] = details.get('fee', Decimal('0'))
                    gz_expenses.append(txn_data)
                    month_stats['gz_expenses'] += amount
                    month_stats['fees'] += txn_data['fee']
                    
                elif txn_type == 'OWNER_PAYMENT':
                    owner_payments.append(txn_data)
                    month_stats['owner_payments'] += amount
                    
                elif txn_type == 'GZ_PAYMENT':
                    gz_payments.append(txn_data)
                    month_stats['gz_payments'] += amount
            
            # 消费记录表
            ws.cell(row=row, column=1, value='📋 消费记录').font = Font(bold=True, size=12)
            row += 1
            create_styled_header(ws, ['日期', '描述/商户', '金额 (RM)', '类型', 'Supplier', '1% Fee'], row, 'FF007F')
            row += 1
            
            for exp in owner_expenses:
                ws.cell(row=row, column=1, value=exp['date'])
                ws.cell(row=row, column=2, value=exp['description'])
                ws.cell(row=row, column=3, value=float(exp['amount'])).number_format = '#,##0.00'
                ws.cell(row=row, column=4, value='Owner Expenses')
                row += 1
            
            for exp in gz_expenses:
                ws.cell(row=row, column=1, value=exp['date'])
                ws.cell(row=row, column=2, value=exp['description'])
                ws.cell(row=row, column=3, value=float(exp['amount'])).number_format = '#,##0.00'
                ws.cell(row=row, column=4, value='GZ Expenses - Supplier')
                ws.cell(row=row, column=5, value=exp['supplier'])
                ws.cell(row=row, column=6, value=float(exp['fee'])).number_format = '#,##0.00'
                row += 1
            
            row += 1
            
            # 付款记录表
            ws.cell(row=row, column=1, value='💳 付款记录').font = Font(bold=True, size=12)
            row += 1
            create_styled_header(ws, ['日期', '描述', '金额 (RM)', '付款方式'], row, 'FF007F')
            row += 1
            
            for pay in owner_payments:
                ws.cell(row=row, column=1, value=pay['date'])
                ws.cell(row=row, column=2, value=pay['description'])
                ws.cell(row=row, column=3, value=float(pay['amount'])).number_format = '#,##0.00'
                ws.cell(row=row, column=4, value='Owner Payment')
                row += 1
            
            for pay in gz_payments:
                ws.cell(row=row, column=1, value=pay['date'])
                ws.cell(row=row, column=2, value=pay['description'])
                ws.cell(row=row, column=3, value=float(pay['amount'])).number_format = '#,##0.00'
                ws.cell(row=row, column=4, value='GZ Direct Payment')
                row += 1
            
            row += 2
            
            # 月度汇总
            ws.cell(row=row, column=1, value='📊 月度汇总').font = Font(bold=True, size=12)
            row += 1
            
            summary_data = [
                ['本月Owner消费总额:', float(month_stats['owner_expenses'])],
                ['本月GZ Supplier消费总额:', float(month_stats['gz_expenses'])],
                ['本月Supplier 1% Fee:', float(month_stats['fees'])],
                ['本月Owner付款总额:', float(month_stats['owner_payments'])],
                ['本月GZ付款总额:', float(month_stats['gz_payments'])],
            ]
            
            for label, value in summary_data:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = 'RM #,##0.00'
                row += 1
            
            auto_adjust_column_width(ws)
            
            # 添加到汇总表
            summary_ws.cell(row=summary_row, column=1, value=bank)
            summary_ws.cell(row=summary_row, column=2, value=month)
            summary_ws.cell(row=summary_row, column=3, value=float(month_stats['owner_expenses'])).number_format = '#,##0.00'
            summary_ws.cell(row=summary_row, column=4, value=float(month_stats['gz_expenses'])).number_format = '#,##0.00'
            summary_ws.cell(row=summary_row, column=5, value=float(month_stats['fees'])).number_format = '#,##0.00'
            summary_ws.cell(row=summary_row, column=6, value=float(month_stats['owner_payments'])).number_format = '#,##0.00'
            summary_ws.cell(row=summary_row, column=7, value=float(month_stats['gz_payments'])).number_format = '#,##0.00'
            summary_row += 1
            
            # 累计总统计
            for key in grand_total:
                grand_total[key] += month_stats[key]
    
    # 添加总计行
    summary_row += 1
    summary_ws.cell(row=summary_row, column=1, value='总计').font = Font(bold=True, size=12)
    summary_ws.cell(row=summary_row, column=3, value=float(grand_total['owner_expenses'])).number_format = '#,##0.00'
    summary_ws.cell(row=summary_row, column=3, value=float(grand_total['owner_expenses'])).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=4, value=float(grand_total['gz_expenses'])).number_format = '#,##0.00'
    summary_ws.cell(row=summary_row, column=4, value=float(grand_total['gz_expenses'])).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=5, value=float(grand_total['fees'])).number_format = '#,##0.00'
    summary_ws.cell(row=summary_row, column=5, value=float(grand_total['fees'])).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=6, value=float(grand_total['owner_payments'])).number_format = '#,##0.00'
    summary_ws.cell(row=summary_row, column=6, value=float(grand_total['owner_payments'])).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=7, value=float(grand_total['gz_payments'])).number_format = '#,##0.00'
    summary_ws.cell(row=summary_row, column=7, value=float(grand_total['gz_payments'])).font = Font(bold=True)
    
    # 最终结算
    summary_row += 3
    summary_ws.cell(row=summary_row, column=1, value='🎯 最终结算金额:').font = Font(bold=True, size=14, color='FF007F')
    
    owner_os = grand_total['owner_expenses'] - grand_total['owner_payments']
    gz_os = grand_total['gz_expenses'] - grand_total['gz_payments']
    
    summary_row += 1
    summary_ws.cell(row=summary_row, column=1, value='Owner OS Balance:')
    summary_ws.cell(row=summary_row, column=2, value=float(owner_os)).number_format = 'RM #,##0.00'
    
    summary_row += 1
    summary_ws.cell(row=summary_row, column=1, value='GZ OS Balance:').font = Font(bold=True, size=12)
    cell = summary_ws.cell(row=summary_row, column=2, value=float(gz_os))
    cell.number_format = 'RM #,##0.00'
    cell.font = Font(bold=True, size=12, color='FF007F')
    
    auto_adjust_column_width(summary_ws)
    
    # 保存Excel文件
    report_dir = 'reports/CCC_Detailed_Reports'
    os.makedirs(report_dir, exist_ok=True)
    excel_file = f'{report_dir}/CCC_Detailed_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(excel_file)
    
    print(f"✅ Excel详细报告已生成: {excel_file}")
    print(f"✅ 共处理 {len(transactions)} 笔交易")
    print(f"✅ 创建 {len(wb.sheetnames)} 个工作表")
    print(f"✅ 覆盖 {len(banks)} 家银行")
    print(f"\n📊 最终统计:")
    print(f"   Owner消费总额:    RM {grand_total['owner_expenses']:,.2f}")
    print(f"   GZ Supplier消费:  RM {grand_total['gz_expenses']:,.2f}")
    print(f"   Supplier 1% Fee:  RM {grand_total['fees']:,.2f}")
    print(f"   Owner付款总额:    RM {grand_total['owner_payments']:,.2f}")
    print(f"   GZ付款总额:       RM {grand_total['gz_payments']:,.2f}")
    print(f"\n🎯 最终结算:")
    print(f"   Owner OS Balance: RM {owner_os:,.2f}")
    print(f"   GZ OS Balance:    RM {gz_os:,.2f}")
    
    conn.close()
    return excel_file

if __name__ == '__main__':
    generate_excel_report()
