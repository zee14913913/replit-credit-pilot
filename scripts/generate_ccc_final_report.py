"""
Chang Choon Chow 最终结算报告生成器
基于monthly_statements表的已验证分类数据
"""
import sqlite3
from decimal import Decimal
from datetime import datetime
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("正在安装openpyxl...")
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl'], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_styled_header(ws, headers, row=1, bg_color='322446'):
    """创建样式化的表头"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_final_report():
    """生成最终结算报告（基于monthly_statements表）"""
    conn = sqlite3.connect('db/smart_loan_manager.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 查询所有Chang Choon Chow的月度账单数据
    cursor.execute("""
        SELECT 
            statement_month,
            bank_name,
            owner_expenses,
            owner_payments,
            gz_expenses,
            gz_payments,
            previous_balance_total,
            closing_balance_total
        FROM monthly_statements
        WHERE customer_id = 10
        ORDER BY statement_month, bank_name
    """)
    
    statements = cursor.fetchall()
    
    # 创建Excel工作簿
    wb = Workbook()
    wb.remove(wb.active)
    
    # 创建汇总表
    summary_ws = wb.create_sheet('最终结算报告', 0)
    row = 1
    
    # 标题
    summary_ws.merge_cells(f'A{row}:I{row}')
    title_cell = summary_ws.cell(row=row, column=1, value='Chang Choon Chow 最终结算报告')
    title_cell.font = Font(bold=True, size=18, color='FF007F')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_ws.row_dimensions[row].height = 30
    row += 2
    
    # 客户信息
    info_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    info_data = [
        ['客户代码:', 'Be_rich_CCC'],
        ['客户姓名:', 'Chang Choon Chow'],
        ['报告生成时间:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['月度记录数:', f'{len(statements)}条'],
        ['时间跨度:', '2024-09 至 2025-10'],
    ]
    
    for info in info_data:
        summary_ws.cell(row=row, column=1, value=info[0]).font = Font(bold=True)
        summary_ws.cell(row=row, column=2, value=info[1])
        summary_ws.cell(row=row, column=1).fill = info_fill
        summary_ws.cell(row=row, column=2).fill = info_fill
        row += 1
    
    row += 2
    
    # 详细数据表头
    create_styled_header(summary_ws, 
                        ['月份', '银行', 'Owner消费', 'Owner付款', 'Owner OS', 'GZ消费', 'GZ付款', 'GZ OS', '月度净增'],
                        row, 'FF007F')
    row += 1
    
    # 总统计
    grand_total = {
        'owner_expenses': Decimal('0'),
        'owner_payments': Decimal('0'),
        'gz_expenses': Decimal('0'),
        'gz_payments': Decimal('0')
    }
    
    owner_os_running = Decimal('0')
    gz_os_running = Decimal('0')
    
    # 逐行添加数据
    for stmt in statements:
        month = stmt['statement_month']
        bank = stmt['bank_name']
        
        owner_exp = abs(Decimal(str(stmt['owner_expenses'])))
        owner_pay = abs(Decimal(str(stmt['owner_payments'])))
        gz_exp = abs(Decimal(str(stmt['gz_expenses'])))
        gz_pay = abs(Decimal(str(stmt['gz_payments'])))
        
        owner_monthly_os = owner_exp - owner_pay
        gz_monthly_os = gz_exp - gz_pay
        monthly_net = owner_monthly_os + gz_monthly_os
        
        owner_os_running += owner_monthly_os
        gz_os_running += gz_monthly_os
        
        summary_ws.cell(row=row, column=1, value=month)
        summary_ws.cell(row=row, column=2, value=bank)
        summary_ws.cell(row=row, column=3, value=float(owner_exp)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=4, value=float(owner_pay)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=5, value=float(owner_monthly_os)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=6, value=float(gz_exp)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=7, value=float(gz_pay)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=8, value=float(gz_monthly_os)).number_format = '#,##0.00'
        summary_ws.cell(row=row, column=9, value=float(monthly_net)).number_format = '#,##0.00'
        
        # 累计
        grand_total['owner_expenses'] += owner_exp
        grand_total['owner_payments'] += owner_pay
        grand_total['gz_expenses'] += gz_exp
        grand_total['gz_payments'] += gz_pay
        
        row += 1
    
    # 总计行
    row += 1
    summary_ws.cell(row=row, column=1, value='总计').font = Font(bold=True, size=12)
    summary_ws.cell(row=row, column=2, value='所有银行').font = Font(bold=True)
    
    total_fill = PatternFill(start_color='FFE6F0', end_color='FFE6F0', fill_type='solid')
    
    for col in range(1, 10):
        summary_ws.cell(row=row, column=col).fill = total_fill
        summary_ws.cell(row=row, column=col).font = Font(bold=True)
    
    summary_ws.cell(row=row, column=3, value=float(grand_total['owner_expenses'])).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=4, value=float(grand_total['owner_payments'])).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=5, value=float(owner_os_running)).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=6, value=float(grand_total['gz_expenses'])).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=7, value=float(grand_total['gz_payments'])).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=8, value=float(gz_os_running)).number_format = '#,##0.00'
    summary_ws.cell(row=row, column=9, value=float(owner_os_running + gz_os_running)).number_format = '#,##0.00'
    
    row += 3
    
    # 最终结算区域
    settlement_fill = PatternFill(start_color='322446', end_color='322446', fill_type='solid')
    
    summary_ws.merge_cells(f'A{row}:I{row}')
    settlement_title = summary_ws.cell(row=row, column=1, value='🎯 最终结算金额')
    settlement_title.font = Font(bold=True, size=16, color='FFFFFF')
    settlement_title.alignment = Alignment(horizontal='center')
    settlement_title.fill = settlement_fill
    summary_ws.row_dimensions[row].height = 25
    row += 2
    
    # Owner OS Balance
    summary_ws.cell(row=row, column=1, value='Owner账本:').font = Font(bold=True, size=12)
    summary_ws.merge_cells(f'B{row}:C{row}')
    row += 1
    
    summary_ws.cell(row=row, column=2, value='总消费:')
    summary_ws.cell(row=row, column=3, value=float(grand_total['owner_expenses'])).number_format = 'RM #,##0.00'
    row += 1
    
    summary_ws.cell(row=row, column=2, value='总付款:')
    summary_ws.cell(row=row, column=3, value=float(grand_total['owner_payments'])).number_format = 'RM #,##0.00'
    row += 1
    
    summary_ws.cell(row=row, column=2, value='Owner OS Balance:').font = Font(bold=True)
    owner_os_cell = summary_ws.cell(row=row, column=3, value=float(owner_os_running))
    owner_os_cell.number_format = 'RM #,##0.00'
    owner_os_cell.font = Font(bold=True, size=12, color='FF007F')
    row += 2
    
    # GZ OS Balance
    summary_ws.cell(row=row, column=1, value='GZ账本:').font = Font(bold=True, size=12)
    row += 1
    
    summary_ws.cell(row=row, column=2, value='GZ总消费 (Supplier):')
    summary_ws.cell(row=row, column=3, value=float(grand_total['gz_expenses'])).number_format = 'RM #,##0.00'
    row += 1
    
    summary_ws.cell(row=row, column=2, value='GZ总付款:')
    summary_ws.cell(row=row, column=3, value=float(grand_total['gz_payments'])).number_format = 'RM #,##0.00'
    row += 1
    
    summary_ws.cell(row=row, column=2, value='GZ OS Balance:').font = Font(bold=True)
    gz_os_cell = summary_ws.cell(row=row, column=3, value=float(gz_os_running))
    gz_os_cell.number_format = 'RM #,##0.00'
    gz_os_cell.font = Font(bold=True, size=14, color='FF007F')
    row += 3
    
    # 最终结算金额（大字突出）
    summary_ws.merge_cells(f'A{row}:C{row}')
    final_settlement_fill = PatternFill(start_color='FF007F', end_color='FF007F', fill_type='solid')
    final_label = summary_ws.cell(row=row, column=1, value='INFINITE GZ 应支付给 OWNER:')
    final_label.font = Font(bold=True, size=14, color='FFFFFF')
    final_label.alignment = Alignment(horizontal='center')
    final_label.fill = final_settlement_fill
    summary_ws.row_dimensions[row].height = 30
    row += 1
    
    summary_ws.merge_cells(f'A{row}:C{row}')
    final_amount = summary_ws.cell(row=row, column=1, value=float(gz_os_running))
    final_amount.number_format = 'RM #,##0.00'
    final_amount.font = Font(bold=True, size=24, color='FF007F')
    final_amount.alignment = Alignment(horizontal='center')
    summary_ws.row_dimensions[row].height = 40
    
    auto_adjust_column_width(summary_ws)
    
    # 创建按银行分组的工作表
    from collections import defaultdict
    banks_data = defaultdict(list)
    
    for stmt in statements:
        banks_data[stmt['bank_name']].append(stmt)
    
    for bank_name, bank_stmts in sorted(banks_data.items()):
        ws = wb.create_sheet(f"{bank_name[:20]}")
        row = 1
        
        # 银行标题
        ws.merge_cells(f'A{row}:H{row}')
        bank_title = ws.cell(row=row, column=1, value=f'{bank_name} - 月度明细')
        bank_title.font = Font(bold=True, size=14, color='FF007F')
        bank_title.alignment = Alignment(horizontal='center')
        row += 2
        
        # 表头
        create_styled_header(ws,
                           ['月份', 'Owner消费', 'Owner付款', 'Owner OS', 'GZ消费', 'GZ付款', 'GZ OS', '月度净增'],
                           row, '322446')
        row += 1
        
        bank_total = {
            'owner_expenses': Decimal('0'),
            'owner_payments': Decimal('0'),
            'gz_expenses': Decimal('0'),
            'gz_payments': Decimal('0'),
            'owner_os': Decimal('0'),
            'gz_os': Decimal('0')
        }
        
        for stmt in sorted(bank_stmts, key=lambda x: x['statement_month']):
            owner_exp = abs(Decimal(str(stmt['owner_expenses'])))
            owner_pay = abs(Decimal(str(stmt['owner_payments'])))
            gz_exp = abs(Decimal(str(stmt['gz_expenses'])))
            gz_pay = abs(Decimal(str(stmt['gz_payments'])))
            
            owner_os = owner_exp - owner_pay
            gz_os = gz_exp - gz_pay
            net = owner_os + gz_os
            
            ws.cell(row=row, column=1, value=stmt['statement_month'])
            ws.cell(row=row, column=2, value=float(owner_exp)).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=float(owner_pay)).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=float(owner_os)).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=float(gz_exp)).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=float(gz_pay)).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=float(gz_os)).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=float(net)).number_format = '#,##0.00'
            
            bank_total['owner_expenses'] += owner_exp
            bank_total['owner_payments'] += owner_pay
            bank_total['gz_expenses'] += gz_exp
            bank_total['gz_payments'] += gz_pay
            bank_total['owner_os'] += owner_os
            bank_total['gz_os'] += gz_os
            
            row += 1
        
        # 银行小计
        row += 1
        ws.cell(row=row, column=1, value=f'{bank_name} 小计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(bank_total['owner_expenses'])).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=float(bank_total['owner_payments'])).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(bank_total['owner_os'])).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(bank_total['gz_expenses'])).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=float(bank_total['gz_payments'])).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=float(bank_total['gz_os'])).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=float(bank_total['owner_os'] + bank_total['gz_os'])).number_format = '#,##0.00'
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        auto_adjust_column_width(ws)
    
    # 保存文件
    report_dir = 'reports/CCC_Detailed_Reports'
    os.makedirs(report_dir, exist_ok=True)
    excel_file = f'{report_dir}/CCC_Final_Settlement_Report.xlsx'
    wb.save(excel_file)
    
    print("=" * 100)
    print("✅ Chang Choon Chow 最终结算报告已生成")
    print("=" * 100)
    print(f"📄 报告文件: {excel_file}")
    print(f"📊 数据来源: monthly_statements表（已验证分类）")
    print(f"📅 时间跨度: 2024-09 至 2025-10")
    print(f"🏦 银行数量: {len(banks_data)}家")
    print(f"📋 月度记录: {len(statements)}条")
    print("")
    print("=" * 100)
    print("💰 最终统计")
    print("=" * 100)
    print(f"Owner总消费:    RM {grand_total['owner_expenses']:>15,.2f}")
    print(f"Owner总付款:    RM {grand_total['owner_payments']:>15,.2f}")
    print(f"Owner OS Balance: RM {owner_os_running:>15,.2f}")
    print("")
    print(f"GZ总消费:       RM {grand_total['gz_expenses']:>15,.2f}")
    print(f"GZ总付款:       RM {grand_total['gz_payments']:>15,.2f}")
    print(f"GZ OS Balance:  RM {gz_os_running:>15,.2f}")
    print("=" * 100)
    print("")
    print("🎯 最终结算金额（GZ OS Balance）")
    print("=" * 100)
    print(f"INFINITE GZ 应支付给 OWNER Chang Choon Chow: RM {gz_os_running:,.2f}")
    print("=" * 100)
    
    conn.close()
    return excel_file, gz_os_running

if __name__ == '__main__':
    generate_final_report()
