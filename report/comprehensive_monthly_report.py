"""
综合月度报告生成器
Generate comprehensive monthly reports with customer/GZ breakdown
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db.database import get_db
from validate.advanced_transaction_analyzer import AdvancedTransactionAnalyzer
from datetime import datetime
import os

class ComprehensiveMonthlyReport:
    
    def __init__(self, output_folder='static/uploads/monthly_reports'):
        self.output_folder = output_folder
        os.makedirs(output_folder, exist_ok=True)
        self.analyzer = AdvancedTransactionAnalyzer()
    
    def generate_monthly_report(self, customer_id, month):
        """
        生成月度综合报告（Excel格式）
        
        包含：
        1. 消费明细表（客户 + GZ）
        2. 付款明细表（客户 + GZ）
        3. 余额分析汇总
        4. Merchant Fee汇总
        """
        # 获取月度数据
        monthly_data = self.analyzer.get_monthly_summary(customer_id, month)
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 删除默认sheet
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
        
        # 1. 创建消费明细表
        self._create_debit_sheet(wb, customer_id, month, monthly_data)
        
        # 2. 创建付款明细表
        self._create_credit_sheet(wb, customer_id, month, monthly_data)
        
        # 3. 创建汇总分析表
        self._create_summary_sheet(wb, customer_id, month, monthly_data)
        
        # 4. 创建Merchant Fee表
        self._create_merchant_fee_sheet(wb, customer_id, month, monthly_data)
        
        # 保存文件
        filename = f"Monthly_Report_{customer_id}_{month}.xlsx"
        filepath = os.path.join(self.output_folder, filename)
        wb.save(filepath)
        
        return {
            'filepath': filepath,
            'filename': filename,
            'month': month,
            'customer_id': customer_id
        }
    
    def _create_debit_sheet(self, wb, customer_id, month, monthly_data):
        """创建消费明细表"""
        ws = wb.create_sheet("消费明细 Debit")
        
        # 样式定义
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        customer_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        gz_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = f"消费明细表 - {month}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头
        headers = ['Bank', 'Card Number', '消费日期', '使用人', '用于', '金额 (RM)', '归属']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row = 4
        customer_debit_total = 0
        gz_debit_total = 0
        
        # 填充数据
        for stmt_data in monthly_data['statements']:
            stmt = stmt_data['statement']
            breakdown = self.analyzer.get_detailed_breakdown(stmt['id'])
            
            # 客户消费
            for tx in breakdown['debit']['customer']:
                ws.cell(row, 1, stmt['bank_name'])
                ws.cell(row, 2, "****")  # 隐藏完整卡号
                ws.cell(row, 3, tx['date'])
                ws.cell(row, 4, tx.get('user_name', '客户'))
                ws.cell(row, 5, tx.get('purpose', '-'))
                ws.cell(row, 6, abs(tx['amount']))
                ws.cell(row, 7, 'Customer')
                
                # 应用样式
                for col in range(1, 8):
                    ws.cell(row, col).fill = customer_fill
                    ws.cell(row, col).border = border
                
                customer_debit_total += abs(tx['amount'])
                row += 1
            
            # GZ消费
            for tx in breakdown['debit']['gz']:
                ws.cell(row, 1, stmt['bank_name'])
                ws.cell(row, 2, "****")
                ws.cell(row, 3, tx['date'])
                ws.cell(row, 4, tx.get('user_name', 'INFINITE GZ'))
                ws.cell(row, 5, tx.get('purpose', '-'))
                ws.cell(row, 6, abs(tx['amount']))
                ws.cell(row, 7, 'GZ')
                
                # 应用样式
                for col in range(1, 8):
                    ws.cell(row, col).fill = gz_fill
                    ws.cell(row, col).border = border
                
                gz_debit_total += abs(tx['amount'])
                row += 1
        
        # 添加Previous Balance
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT SUM(customer_previous_balance) as total
                FROM statement_balance_analysis
                WHERE customer_id = ?
                AND strftime('%Y-%m', (SELECT statement_date FROM statements WHERE id = statement_id)) = ?
            ''', (customer_id, month))
            prev_balance = cursor.fetchone()['total'] or 0
        
        # 汇总行
        row += 1
        ws.cell(row, 5, 'Previous Balance:').font = Font(bold=True)
        ws.cell(row, 6, prev_balance).font = Font(bold=True)
        
        row += 1
        ws.cell(row, 5, '客户消费汇总:').font = Font(bold=True)
        ws.cell(row, 6, customer_debit_total).font = Font(bold=True)
        ws.cell(row, 7, 'Customer').font = Font(bold=True)
        
        row += 1
        ws.cell(row, 5, 'GZ消费汇总:').font = Font(bold=True)
        ws.cell(row, 6, gz_debit_total).font = Font(bold=True)
        ws.cell(row, 7, 'GZ').font = Font(bold=True)
        
        row += 1
        ws.cell(row, 5, '总消费 (含Previous):').font = Font(bold=True, color="FF0000")
        ws.cell(row, 6, prev_balance + customer_debit_total + gz_debit_total).font = Font(bold=True, color="FF0000")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
    
    def _create_credit_sheet(self, wb, customer_id, month, monthly_data):
        """创建付款明细表"""
        ws = wb.create_sheet("付款明细 Credit")
        
        # 样式
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        customer_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        gz_fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = f"付款明细表 - {month}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头
        headers = ['Bank', 'Card Number', '付款日期', '付款人', '付款于', '金额 (RM)', '归属']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row = 4
        customer_credit_total = 0
        gz_credit_total = 0
        
        # 填充数据
        for stmt_data in monthly_data['statements']:
            stmt = stmt_data['statement']
            breakdown = self.analyzer.get_detailed_breakdown(stmt['id'])
            
            # 客户付款
            for tx in breakdown['credit']['customer']:
                ws.cell(row, 1, stmt['bank_name'])
                ws.cell(row, 2, "****")
                ws.cell(row, 3, tx['date'])
                ws.cell(row, 4, tx.get('user_name', '客户本人'))
                ws.cell(row, 5, tx.get('purpose', '还款'))
                ws.cell(row, 6, abs(tx['amount']))
                ws.cell(row, 7, 'Customer')
                
                for col in range(1, 8):
                    ws.cell(row, col).fill = customer_fill
                    ws.cell(row, col).border = border
                
                customer_credit_total += abs(tx['amount'])
                row += 1
            
            # GZ付款
            for tx in breakdown['credit']['gz']:
                ws.cell(row, 1, stmt['bank_name'])
                ws.cell(row, 2, "****")
                ws.cell(row, 3, tx['date'])
                ws.cell(row, 4, tx.get('user_name', 'INFINITE GZ'))
                ws.cell(row, 5, tx.get('purpose', '代付'))
                ws.cell(row, 6, abs(tx['amount']))
                ws.cell(row, 7, 'GZ')
                
                for col in range(1, 8):
                    ws.cell(row, col).fill = gz_fill
                    ws.cell(row, col).border = border
                
                gz_credit_total += abs(tx['amount'])
                row += 1
        
        # 汇总
        row += 1
        ws.cell(row, 5, '客户付款汇总:').font = Font(bold=True)
        ws.cell(row, 6, customer_credit_total).font = Font(bold=True)
        ws.cell(row, 7, 'Customer').font = Font(bold=True)
        
        row += 1
        ws.cell(row, 5, 'GZ付款汇总:').font = Font(bold=True)
        ws.cell(row, 6, gz_credit_total).font = Font(bold=True)
        ws.cell(row, 7, 'GZ').font = Font(bold=True)
        
        row += 1
        ws.cell(row, 5, '总付款:').font = Font(bold=True, color="00AA00")
        ws.cell(row, 6, customer_credit_total + gz_credit_total).font = Font(bold=True, color="00AA00")
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
    
    def _create_summary_sheet(self, wb, customer_id, month, monthly_data):
        """创建汇总分析表"""
        ws = wb.create_sheet("余额分析 Summary", 0)  # 设为第一个sheet
        
        # 获取客户信息
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM customers WHERE id = ?', (customer_id,))
            customer = dict(cursor.fetchone())
        
        summary = monthly_data['summary']
        
        # 标题
        ws.merge_cells('A1:E1')
        ws['A1'] = f"月度余额分析 - {month}"
        ws['A1'].font = Font(bold=True, size=18)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 客户信息
        row = 3
        ws.cell(row, 1, '客户姓名:').font = Font(bold=True)
        ws.cell(row, 2, customer['name'])
        row += 1
        ws.cell(row, 1, 'IC号码:').font = Font(bold=True)
        ws.cell(row, 2, customer.get('ic_number', '-'))
        row += 2
        
        # 客户部分
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row, 1, '【客户 Customer】').font = Font(bold=True, size=14, color="0066CC")
        row += 1
        
        data = [
            ['项目', '金额 (RM)'],
            ['Previous Balance', summary['customer'].get('previous_balance', 0)],
            ['消费总额 (Transactions)', summary['customer']['total_debit']],
            ['付款总额', summary['customer']['total_credit']],
            ['客户余额', summary['customer']['balance']]
        ]
        
        for item in data:
            for col, value in enumerate(item, 1):
                cell = ws.cell(row, col, value)
                if row == data.index(item) + row - len(data) + 1:
                    cell.font = Font(bold=True)
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        
        row += 1
        
        # GZ部分
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row, 1, '【INFINITE GZ】').font = Font(bold=True, size=14, color="FF6600")
        row += 1
        
        gz_data = [
            ['项目', '金额 (RM)'],
            ['消费总额', summary['gz']['total_debit']],
            ['付款总额', summary['gz']['total_credit']],
            ['GZ余额', summary['gz']['balance']]
        ]
        
        for item in gz_data:
            for col, value in enumerate(item, 1):
                cell = ws.cell(row, col, value)
                if row == gz_data.index(item) + row - len(gz_data) + 1:
                    cell.font = Font(bold=True)
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Merchant Fee
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row, 1, '【手续费 Merchant Fee】').font = Font(bold=True, size=14, color="CC0000")
        row += 1
        ws.cell(row, 1, 'Total Merchant Fee (1%):').font = Font(bold=True)
        ws.cell(row, 2, summary['merchant_fee_total']).font = Font(bold=True, color="CC0000")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_merchant_fee_sheet(self, wb, customer_id, month, monthly_data):
        """创建Merchant Fee详细表"""
        ws = wb.create_sheet("Merchant Fee")
        
        # 表头
        ws['A1'] = f"Merchant Fee详细 - {month}"
        ws['A1'].font = Font(bold=True, size=16)
        
        headers = ['Date', 'Supplier', 'Amount', 'Fee (1%)', 'Bank']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
        
        row = 4
        total_fee = 0
        
        for stmt_data in monthly_data['statements']:
            stmt = stmt_data['statement']
            
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT transaction_date, description, amount, supplier_fee
                    FROM transactions
                    WHERE statement_id = ? AND supplier_fee > 0
                    ORDER BY transaction_date
                ''', (stmt['id'],))
                
                for tx in cursor.fetchall():
                    ws.cell(row, 1, tx['transaction_date'])
                    ws.cell(row, 2, tx['description'][:50])  # Supplier name from description
                    ws.cell(row, 3, abs(tx['amount']))
                    ws.cell(row, 4, tx['supplier_fee'])
                    ws.cell(row, 5, stmt['bank_name'])
                    total_fee += tx['supplier_fee']
                    row += 1
        
        # 总计
        row += 1
        ws.cell(row, 3, 'Total Fee:').font = Font(bold=True)
        ws.cell(row, 4, total_fee).font = Font(bold=True, color="CC0000")
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20


# 便捷函数
def generate_report(customer_id, month):
    """生成月度报告"""
    generator = ComprehensiveMonthlyReport()
    return generator.generate_monthly_report(customer_id, month)
