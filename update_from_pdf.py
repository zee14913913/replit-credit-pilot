#!/usr/bin/env python3
"""
从PDF批量更新数据库 - 安全更新脚本
=================================
严格规则：
1. 只更新PDF成功提取的字段
2. 生成详细的更新日志
3. 提供回滚SQL
"""

import sqlite3
import openpyxl
from datetime import datetime

def update_database_from_excel(excel_file: str, dry_run: bool = True):
    """从Excel报告批量更新数据库"""
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    conn = sqlite3.connect('db/smart_loan_manager.db')
    cursor = conn.cursor()
    
    updates = []
    rollback_sql = []
    
    print(f"\n{'='*100}")
    print(f"📊 从Excel报告批量更新数据库")
    print(f"{'='*100}\n")
    print(f"模式: {'🔍 DRY RUN (仅预览，不实际更新)' if dry_run else '✅ LIVE RUN (实际更新数据库)'}\n")
    
    row_num = 2
    update_count = 0
    
    while row_num <= ws.max_row:
        stmt_id = ws.cell(row=row_num, column=1).value
        status = ws.cell(row=row_num, column=13).value
        
        # 只处理有差异或成功提取的记录
        if stmt_id and status and ('差异' in str(status) or '正确' in str(status)):
            # 获取PDF值
            pdf_stmt_date = ws.cell(row=row_num, column=6).value
            pdf_due_date = ws.cell(row=row_num, column=8).value
            pdf_total = ws.cell(row=row_num, column=10).value
            pdf_min_pay = ws.cell(row=row_num, column=12).value
            
            # 获取数据库当前值（用于回滚）
            cursor.execute("""
                SELECT statement_date, due_date, statement_total, minimum_payment
                FROM statements
                WHERE id = ?
            """, (stmt_id,))
            
            old_values = cursor.fetchone()
            if not old_values:
                row_num += 1
                continue
            
            old_stmt_date, old_due_date, old_total, old_min_pay = old_values
            
            # 准备更新语句
            updates_for_this_record = []
            set_clauses = []
            values = []
            
            # Statement Date
            if pdf_stmt_date and pdf_stmt_date != old_stmt_date:
                set_clauses.append("statement_date = ?")
                values.append(pdf_stmt_date)
                updates_for_this_record.append(f"Statement Date: {old_stmt_date} → {pdf_stmt_date}")
            
            # Due Date
            if pdf_due_date and pdf_due_date != old_due_date:
                set_clauses.append("due_date = ?")
                values.append(pdf_due_date)
                updates_for_this_record.append(f"Due Date: {old_due_date} → {pdf_due_date}")
            
            # Statement Total
            if pdf_total and (not old_total or abs(float(old_total) - float(pdf_total)) > 0.01):
                set_clauses.append("statement_total = ?")
                values.append(float(pdf_total))
                updates_for_this_record.append(f"Statement Total: RM {old_total or 0:.2f} → RM {pdf_total:.2f}")
            
            # Minimum Payment
            if pdf_min_pay and (not old_min_pay or abs(float(old_min_pay) - float(pdf_min_pay)) > 0.01):
                set_clauses.append("minimum_payment = ?")
                values.append(float(pdf_min_pay))
                updates_for_this_record.append(f"Minimum Payment: RM {old_min_pay or 0:.2f} → RM {pdf_min_pay:.2f}")
            
            # 如果有需要更新的字段
            if set_clauses:
                values.append(stmt_id)
                update_sql = f"UPDATE statements SET {', '.join(set_clauses)} WHERE id = ?"
                
                # 生成回滚SQL
                rollback_values = []
                rollback_clauses = []
                if "statement_date" in update_sql:
                    rollback_clauses.append("statement_date = ?")
                    rollback_clauses.append(old_stmt_date if old_stmt_date else "NULL")
                if "due_date" in update_sql:
                    rollback_clauses.append("due_date = ?")
                    rollback_values.append(old_due_date if old_due_date else "NULL")
                if "statement_total" in update_sql:
                    rollback_clauses.append("statement_total = ?")
                    rollback_values.append(old_total if old_total else "NULL")
                if "minimum_payment" in update_sql:
                    rollback_clauses.append("minimum_payment = ?")
                    rollback_values.append(old_min_pay if old_min_pay else "NULL")
                
                rollback_sql.append(f"-- Rollback for Statement ID {stmt_id}")
                rollback_sql.append(f"UPDATE statements SET {', '.join(rollback_clauses)} WHERE id = {stmt_id};")
                
                # 打印更新信息
                customer = ws.cell(row=row_num, column=2).value
                bank = ws.cell(row=row_num, column=3).value
                
                print(f"✅ Statement ID {stmt_id}: {customer} - {bank}")
                for update_desc in updates_for_this_record:
                    print(f"   {update_desc}")
                print()
                
                updates.append((update_sql, values))
                update_count += 1
        
        row_num += 1
    
    # 执行更新
    if not dry_run and updates:
        for update_sql, values in updates:
            cursor.execute(update_sql, values)
        
        conn.commit()
        print(f"\n✅ 成功更新 {update_count} 条记录到数据库！\n")
    elif dry_run:
        print(f"\n🔍 DRY RUN完成：发现 {update_count} 条需要更新的记录")
        print(f"   运行脚本时使用 dry_run=False 参数来实际更新数据库\n")
    else:
        print(f"\n⚠️  没有需要更新的记录\n")
    
    # 保存回滚SQL
    if rollback_sql:
        with open('rollback_updates.sql', 'w') as f:
            f.write("-- 回滚SQL脚本\n")
            f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"-- 总计: {update_count} 条记录\n\n")
            f.write('\n'.join(rollback_sql))
        
        print(f"📁 回滚SQL已保存到: rollback_updates.sql\n")
    
    conn.close()
    
    print(f"{'='*100}\n")

if __name__ == "__main__":
    # 先进行DRY RUN预览
    print("\n第一步：DRY RUN - 预览更新\n")
    update_database_from_excel('pdf_vs_database_comparison.xlsx', dry_run=True)
    
    # 询问是否继续
    print("\n" + "="*100)
    print("⚠️  确认更新")
    print("="*100)
    print("\n请检查上面的更新内容。如果确认无误，请手动运行：")
    print("\n  python3 update_from_pdf.py --confirm\n")
    print("这将实际更新数据库。\n")
    
    # 检查命令行参数
    import sys
    if '--confirm' in sys.argv:
        print("\n开始实际更新...\n")
        update_database_from_excel('pdf_vs_database_comparison.xlsx', dry_run=False)
